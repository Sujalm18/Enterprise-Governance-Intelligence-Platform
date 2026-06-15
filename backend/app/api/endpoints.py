import logging
import os
import re
import shutil
import time
from datetime import datetime
from pathlib import Path
from typing import List, Optional
from fastapi import APIRouter, Depends, UploadFile, File, BackgroundTasks, HTTPException, Query, status, Header
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from sqlalchemy.exc import OperationalError

from backend.app.config import settings
from backend.app.database import get_db
from backend.app.migrations import is_schema_mismatch_error, schema_mismatch_response_detail
from backend.app.models import (
    Document, WorkflowJob, WorkflowStatus, GovernanceReport, RaidItem, EscalationItem, AuditLog, User, MitigationTask, Notification, MeetingAction, Organization, GovernanceTrendSnapshot
)
from backend.app.schemas import (
    DocumentResponse, WorkflowJobResponse, GovernanceReportResponse,
    ReportReviewRequest, EscalationRouteRequest, DashboardStatsResponse,
    EscalationItemResponse, AuditLogResponse, DashboardChartsResponse,
    ReportAssignRequest, EscalationAssignRequest, MitigationTaskResponse,
    MitigationTaskUpdateRequest, NotificationResponse, NotificationReadRequest,
    DemoDataGenerateRequest, InboxResponse,
    GovernanceMaturityResponse, ExecutivePriorityItem, RootCauseAnalyticsResponse,
    StrategicRecommendationsResponse, ExecutiveBriefingResponse, CopilotRequest,
    CopilotResponse, GovernanceTrendsResponse, HealthExplanationsResponse,
    HealthExplanationItem, LoginRequest, TokenResponse
)
from backend.app.services.workflow import process_document_pipeline

logger = logging.getLogger("governance_copilot.api.endpoints")
router = APIRouter()


from backend.app.auth import (
    get_current_role,
    get_current_tenant,
    verify_password,
    create_access_token,
    get_current_user_optional
)


@router.post("/auth/login", response_model=TokenResponse)
def login(login_data: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == login_data.username).first()
    if not user or not user.password_hash or not verify_password(login_data.password, user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token = create_access_token(data={"sub": user.username})
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "role": user.role,
        "tenant_id": user.tenant_id or 1
    }


@router.get("/auth/me")
def get_me(current_user: Optional[User] = Depends(get_current_user_optional)):
    if not current_user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Not authenticated"
        )
    return {
        "id": current_user.id,
        "username": current_user.username,
        "role": current_user.role,
        "tenant_id": current_user.tenant_id or 1
    }



def _raise_schema_mismatch_http_error(error: OperationalError) -> None:
    logger.error("Database schema mismatch detected while querying reports: %s", error)
    raise HTTPException(
        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        detail=schema_mismatch_response_detail(error)
    )

@router.post("/upload", response_model=DocumentResponse, status_code=status.HTTP_201_CREATED)
async def upload_document(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    chunk_size: Optional[int] = Query(None),
    chunk_overlap: Optional[int] = Query(None),
    use_rag: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role)
):
    """
    Uploads a document, creates a DB entry and WorkflowJob,
    and runs the background worker pipeline.
    """
    logger.info(f"Incoming upload request for file: {file.filename} by role: {x_user_role}")
    
    # Validate extension
    file_ext = file.filename.split(".")[-1].lower() if "." in file.filename else ""
    if file_ext not in ("pdf", "docx", "txt", "doc"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported file format '.{file_ext}'. Supported formats: PDF, DOCX, TXT"
        )
        
    # Override settings if provided in request
    if chunk_size is not None:
        settings.CHUNK_SIZE = chunk_size
    if chunk_overlap is not None:
        settings.CHUNK_OVERLAP = chunk_overlap
    if use_rag is not None:
        settings.USE_RAG = use_rag

    # Save file to uploads folder
    filename_clean = f"{int(time.time())}_{file.filename}"
    file_path = Path(settings.UPLOAD_DIR) / filename_clean
    
    try:
        with open(file_path, "wb") as f:
            shutil.copyfileobj(file.file, f)
    except Exception as e:
        logger.error(f"Failed to write file to disk: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Could not save file to disk."
        )

    # 1. Create Document Record
    doc = Document(
        filename=str(file_path),
        type=file_ext,
        status=WorkflowStatus.UPLOADED
    )
    db.add(doc)
    db.commit()
    
    # 2. Create WorkflowJob Record
    job = WorkflowJob(
        document_id=doc.id,
        status=WorkflowStatus.UPLOADED,
        logs=f"File uploaded. Saved to: {file_path.name}\n"
    )
    db.add(job)
    db.commit()
    
    # 3. Log Audit Trail (Unified AuditEvent format)
    audit = AuditLog(
        document_id=doc.id,
        event="Uploaded",
        user="analyst_user",
        user_role=x_user_role,
        action="document upload",
        entity_type="document",
        entity_id=doc.id,
        details=f"Document '{file.filename}' uploaded and queued for processing."
    )
    db.add(audit)
    db.commit()

    # 4. Dispatch background worker pipeline task
    background_tasks.add_task(
        process_document_pipeline,
        document_id=doc.id,
        job_id=job.id
    )

    logger.info(f"Background task queued. Job ID: {job.id}, Document ID: {doc.id}")
    return doc

@router.get("/workflow/jobs/{id}", response_model=WorkflowJobResponse)
def get_workflow_job(id: int, db: Session = Depends(get_db)):
    """Retrieves status and logs of a workflow job."""
    job = db.query(WorkflowJob).filter(WorkflowJob.id == id).first()
    if not job:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Workflow job not found")
    return job

@router.get("/governance/reports", response_model=List[GovernanceReportResponse])
def list_governance_reports(
    is_latest: bool = Query(True),
    review_status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    """Lists governance reports, default returns latest version only."""
    # Eager-load Document relationship to avoid N+1 queries and include filename
    query = db.query(GovernanceReport).options(joinedload(GovernanceReport.document)).filter(GovernanceReport.tenant_id == tenant_id)
    if is_latest:
        query = query.filter(GovernanceReport.is_latest == True)
    if review_status:
        query = query.filter(GovernanceReport.review_status == review_status)

    try:
        reports = query.order_by(GovernanceReport.created_at.desc()).all()
    except OperationalError as error:
        if is_schema_mismatch_error(error):
            _raise_schema_mismatch_http_error(error)
        raise
    return reports

@router.get("/governance/reports/{id}", response_model=GovernanceReportResponse)
def get_governance_report(id: int, db: Session = Depends(get_db)):
    """Retrieves a single governance report by ID with nested details."""
    try:
        report = db.query(GovernanceReport).options(joinedload(GovernanceReport.document)).filter(GovernanceReport.id == id).first()
    except OperationalError as error:
        if is_schema_mismatch_error(error):
            _raise_schema_mismatch_http_error(error)
        raise
    if not report:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Governance report not found")
    return report

@router.patch("/governance/reports/{id}/review", response_model=GovernanceReportResponse)
def review_governance_report(
    id: int,
    payload: ReportReviewRequest,
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role)
):
    """Approves or requests changes for a report, driving the workflow job state."""
    if x_user_role != "Manager":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Managers can review reports."
        )
    try:
        report = db.query(GovernanceReport).filter(GovernanceReport.id == id).first()
    except OperationalError as error:
        if is_schema_mismatch_error(error):
            _raise_schema_mismatch_http_error(error)
        raise
    if not report:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Governance report not found")
        
    # Update report fields
    report.reviewer = payload.reviewer
    report.review_status = payload.review_status
    report.review_notes = payload.review_notes
    report.updated_at = datetime.utcnow()
    
    # Drive Document/Job state based on action
    doc = db.query(Document).filter(Document.id == report.document_id).first()
    job = db.query(WorkflowJob).filter(WorkflowJob.document_id == report.document_id).order_by(WorkflowJob.id.desc()).first()
    
    audit_details = f"Review Notes: {payload.review_notes or 'None'}"
    
    if payload.review_status == "approved":
        # Report approved -> Published
        doc.status = WorkflowStatus.PUBLISHED
        if job:
            job.status = WorkflowStatus.PUBLISHED
        report.review_status = "approved"
        report.status = "APPROVED"
        report.approved_by = payload.reviewer
        event_name = "Approved"
        action_name = "report approval"
    else:
        # Changes requested -> Failed review state
        doc.status = WorkflowStatus.FAILED
        if job:
            job.status = WorkflowStatus.FAILED
        report.review_status = "changes_requested"
        report.status = "DRAFT" # send back to draft
        event_name = "Changes Requested"
        action_name = "report changes requested"
        
    # Record Audit entry
    audit = AuditLog(
        document_id=report.document_id,
        governance_report_id=report.id,
        event=event_name,
        user=payload.reviewer,
        user_role=x_user_role,
        action=action_name,
        entity_type="report",
        entity_id=report.id,
        details=f"Status: {payload.review_status}. {audit_details}"
    )
    
    db.add(audit)
    
    # Create notification for Analyst
    notif_severity = "LOW" if payload.review_status == "approved" else "MEDIUM"
    notif_title = "Governance Report Approved" if payload.review_status == "approved" else "Changes Requested on Governance Report"
    notif_msg = f"Governance Report V{report.version} has been approved by {payload.reviewer}." if payload.review_status == "approved" else f"Governance Report V{report.version} requires changes. Notes: {payload.review_notes or 'None'}."
    
    notif = Notification(
        severity=notif_severity,
        notification_type="GOVERNANCE_ALERT",
        title=notif_title,
        message=notif_msg,
        recipient_role="Analyst",
        related_entity_type="report",
        related_entity_id=report.id,
        read_status=False
    )
    db.add(notif)
    
    db.commit()
    db.refresh(report)
    
    logger.info(f"Report ID {report.id} reviewed. Status updated to {payload.review_status} by {payload.reviewer}")
    return report

def check_dynamic_sla_notifications(db: Session):
    """
    Checks active mitigations target dates pull-style.
    - Due within 3 days -> Generate MITIGATION_DUE_SOON if not exists.
    - Overdue -> Generate MITIGATION_OVERDUE and SLA_BREACH if not exists.
    """
    active_tasks = db.query(MitigationTask).filter(
        MitigationTask.status.notin_(["COMPLETED", "VERIFIED"])
    ).all()
    
    today = datetime.utcnow().date()
    for task in active_tasks:
        if not task.target_date:
            continue
        try:
            due_date = datetime.strptime(task.target_date, "%Y-%m-%d").date()
        except ValueError:
            continue
            
        # 1. Overdue checks
        if due_date < today:
            # Overdue alert
            existing_overdue = db.query(Notification).filter(
                Notification.notification_type == "MITIGATION_OVERDUE",
                Notification.related_entity_id == task.id,
                Notification.recipient_role == task.owner_role
            ).first()
            if not existing_overdue:
                notif_overdue = Notification(
                    severity="HIGH",
                    notification_type="MITIGATION_OVERDUE",
                    title="Mitigation Task Overdue",
                    message=f"Mitigation task '{task.title}' is overdue! Target date was {task.target_date}.",
                    recipient_role=task.owner_role,
                    related_entity_type="mitigation",
                    related_entity_id=task.id,
                    read_status=False,
                    created_at=datetime.utcnow()
                )
                db.add(notif_overdue)
                
            # SLA breach alert for Governance Lead & Manager
            for role in ("Manager", "Governance Lead"):
                existing_breach = db.query(Notification).filter(
                    Notification.notification_type == "SLA_BREACH",
                    Notification.related_entity_id == task.id,
                    Notification.recipient_role == role
                ).first()
                if not existing_breach:
                    notif_breach = Notification(
                        severity="CRITICAL",
                        notification_type="SLA_BREACH",
                        title="Mitigation SLA Breach",
                        message=f"SLA Breach: Mitigation task '{task.title}' owned by {task.owner_role} is overdue since {task.target_date}.",
                        recipient_role=role,
                        related_entity_type="mitigation",
                        related_entity_id=task.id,
                        read_status=False,
                        created_at=datetime.utcnow()
                    )
                    db.add(notif_breach)
                    if role == "Manager":
                        try:
                            from backend.app.services.integrations import trigger_governance_alerts
                            trigger_governance_alerts(
                                db=db,
                                tenant_id=task.tenant_id or 1,
                                title="SLA Breach: Overdue Mitigation Task",
                                message=f"SLA Breach: Mitigation task '{task.title}' owned by {task.owner_role} is overdue since {task.target_date}.",
                                severity="critical",
                                details=f"Task Owner: {task.owner_name or 'Unassigned'} ({task.owner_role})\nTarget Date: {task.target_date}"
                            )
                        except Exception as alert_err:
                            logger.error(f"Failed to trigger SLA breach alert webhook: {alert_err}")
                    
        # 2. Due soon checks (<= 3 days left, and NOT overdue)
        elif (due_date - today).days <= 3:
            existing_duesoon = db.query(Notification).filter(
                Notification.notification_type == "MITIGATION_DUE_SOON",
                Notification.related_entity_id == task.id,
                Notification.recipient_role == task.owner_role
            ).first()
            if not existing_duesoon:
                notif_duesoon = Notification(
                    severity="MEDIUM",
                    notification_type="MITIGATION_DUE_SOON",
                    title="Mitigation Task Due Soon",
                    message=f"Mitigation task '{task.title}' is due in {(due_date - today).days} days.",
                    recipient_role=task.owner_role,
                    related_entity_type="mitigation",
                    related_entity_id=task.id,
                    read_status=False,
                    created_at=datetime.utcnow()
                )
                db.add(notif_duesoon)
                
    db.commit()


@router.get("/governance/dashboard/stats", response_model=DashboardStatsResponse)
def get_dashboard_stats(
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role),
    tenant_id: int = Depends(get_current_tenant)
):
    """Computes executive KPIs across the entire governance platform."""
    # 0. Sync dynamic SLA alerts
    check_dynamic_sla_notifications(db)

    # Count total processed documents
    total_docs = db.query(Document).filter(Document.tenant_id == tenant_id).count()
    
    # Reports pending manager review
    pending_reviews = db.query(GovernanceReport).filter(
        GovernanceReport.tenant_id == tenant_id,
        GovernanceReport.status == "PENDING_MANAGER_REVIEW"
    ).count()
    
    # Total approved reports
    approved_reports = db.query(GovernanceReport).filter(
        GovernanceReport.tenant_id == tenant_id,
        GovernanceReport.review_status == "approved"
    ).count()
    
    # Failed processing jobs
    failed_jobs = db.query(WorkflowJob).filter(
        WorkflowJob.tenant_id == tenant_id,
        WorkflowJob.status == WorkflowStatus.FAILED
    ).count()
    
    # Escalations stats
    total_escalations = db.query(EscalationItem).filter(EscalationItem.tenant_id == tenant_id).count()
    open_escalations = db.query(EscalationItem).filter(
        EscalationItem.tenant_id == tenant_id,
        EscalationItem.status.notin_(["RESOLVED", "CLOSED"])
    ).count()

    # Average confidence score across all reports
    avg_confidence = db.query(func.avg(GovernanceReport.confidence_score)).filter(GovernanceReport.tenant_id == tenant_id).scalar() or 1.0
    
    # Average processing time
    avg_processing = db.query(func.avg(GovernanceReport.processing_time_seconds)).filter(GovernanceReport.tenant_id == tenant_id).scalar() or 0.0
    
    # Total tokens consumed
    total_tokens = db.query(func.sum(GovernanceReport.tokens_used)).filter(GovernanceReport.tenant_id == tenant_id).scalar() or 0
    
    # Reports generated count
    reports_generated = db.query(GovernanceReport).filter(GovernanceReport.tenant_id == tenant_id).count()
    
    # Recent audit logs
    recent_audits = db.query(AuditLog).filter(AuditLog.tenant_id == tenant_id).order_by(AuditLog.timestamp.desc()).limit(10).all()

    # Phase 3 Mitigation KPIs
    # 1. Total original risk (sum of risk_score for type in ('risk', 'issue'))
    total_original_risk = db.query(func.sum(RaidItem.risk_score)).filter(
        RaidItem.tenant_id == tenant_id,
        RaidItem.type.in_(["risk", "issue"])
    ).scalar() or 0

    # 2. Total current risk (sum of current_risk_score for type in ('risk', 'issue'))
    total_current_risk = db.query(func.sum(RaidItem.current_risk_score)).filter(
        RaidItem.tenant_id == tenant_id,
        RaidItem.type.in_(["risk", "issue"])
    ).scalar() or 0

    # 3. Risk reduction percentage
    if total_original_risk > 0:
        risk_reduction_percentage = round(((total_original_risk - total_current_risk) / total_original_risk) * 100.0, 1)
    else:
        risk_reduction_percentage = 0.0

    # 4. Overdue mitigations count & dynamic SLA calculations
    active_tasks = db.query(MitigationTask).filter(
        MitigationTask.tenant_id == tenant_id,
        MitigationTask.status.notin_(["COMPLETED", "VERIFIED"])
    ).all()
    
    overdue_mitigations_count = 0
    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    for task in active_tasks:
        if task.target_date and task.target_date < today_str:
            overdue_mitigations_count += 1

    # 5. Governance Health Score calculation
    # Start at 100
    governance_health_score = 100
    
    # Deduct 5 points for each open critical or high RAID item (current_risk_score > 0 and severity in ('critical', 'high'))
    open_risks_count = db.query(RaidItem).filter(
        RaidItem.tenant_id == tenant_id,
        RaidItem.type.in_(["risk", "issue"]),
        RaidItem.severity.in_(["critical", "high"]),
        RaidItem.current_risk_score > 0
    ).count()
    governance_health_score -= open_risks_count * 5

    # Deduct 8 points for each open EscalationItem (status not in ('RESOLVED', 'CLOSED'))
    governance_health_score -= open_escalations * 8

    # Deduct 4 points for each OVERDUE mitigation task
    governance_health_score -= overdue_mitigations_count * 4

    # Add 2 bonus points for each VERIFIED mitigation task (capped at maximum +15 bonus)
    verified_count = db.query(MitigationTask).filter(
        MitigationTask.tenant_id == tenant_id,
        MitigationTask.status == "VERIFIED"
    ).count()
    bonus = min(verified_count * 2, 15)
    governance_health_score += bonus

    # Bounded strictly between 0 and 100
    governance_health_score = max(0, min(100, governance_health_score))

    # 6. Mitigations pipeline counts
    mitigations_pipeline_counts = {"PLANNED": 0, "IN_PROGRESS": 0, "BLOCKED": 0, "COMPLETED": 0, "VERIFIED": 0}
    status_counts = db.query(MitigationTask.status, func.count(MitigationTask.id)).filter(MitigationTask.tenant_id == tenant_id).group_by(MitigationTask.status).all()
    for status_val, count in status_counts:
        if status_val in mitigations_pipeline_counts:
            mitigations_pipeline_counts[status_val] = count

    # Phase 4 & 4.5 unread, breaches, approvals
    unread_notifications = db.query(Notification).filter(
        Notification.tenant_id == tenant_id,
        Notification.recipient_role == x_user_role,
        Notification.read_status == False
    ).count()

    sla_breaches_count = overdue_mitigations_count

    pending_governance_approvals = db.query(MitigationTask).filter(
        MitigationTask.tenant_id == tenant_id,
        MitigationTask.status == "COMPLETED"
    ).count()

    return {
        "total_documents": total_docs,
        "pending_reviews": pending_reviews,
        "approved_reports": approved_reports,
        "failed_jobs": failed_jobs,
        "total_escalations": total_escalations,
        "open_escalations": open_escalations,
        "average_confidence": round(float(avg_confidence), 2),
        "average_processing_time": round(float(avg_processing), 2),
        "total_tokens_consumed": int(total_tokens),
        "reports_generated": reports_generated,
        "recent_logs": recent_audits,
        "governance_health_score": governance_health_score,
        "total_original_risk": total_original_risk,
        "total_current_risk": total_current_risk,
        "risk_reduction_percentage": risk_reduction_percentage,
        "overdue_mitigations_count": overdue_mitigations_count,
        "mitigations_pipeline_counts": mitigations_pipeline_counts,
        "unread_notifications": unread_notifications,
        "sla_breaches_count": sla_breaches_count,
        "pending_governance_approvals": pending_governance_approvals
    }


@router.get("/governance/dashboard/charts", response_model=DashboardChartsResponse)
def get_dashboard_charts(
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    """Returns aggregated chart data for the executive dashboard."""
    
    # Reports by review status
    reports_by_status_rows = (
        db.query(GovernanceReport.review_status, func.count(GovernanceReport.id))
        .filter(GovernanceReport.tenant_id == tenant_id)
        .group_by(GovernanceReport.review_status)
        .all()
    )
    reports_by_status = [
        {"label": status or "unknown", "count": count}
        for status, count in reports_by_status_rows
    ]
    
    # Escalations by severity
    esc_by_severity_rows = (
        db.query(EscalationItem.severity, func.count(EscalationItem.id))
        .filter(EscalationItem.tenant_id == tenant_id)
        .group_by(EscalationItem.severity)
        .all()
    )
    escalations_by_severity = [
        {"label": severity or "unknown", "count": count}
        for severity, count in esc_by_severity_rows
    ]
    
    # RAID items distribution by type
    raid_dist_rows = (
        db.query(RaidItem.type, func.count(RaidItem.id))
        .filter(RaidItem.tenant_id == tenant_id)
        .group_by(RaidItem.type)
        .all()
    )
    raid_distribution = [
        {"label": raid_type or "unknown", "count": count}
        for raid_type, count in raid_dist_rows
    ]
    
    # Processing trend: reports created per day (last 30 days)
    trend_rows = (
        db.query(
            func.date(GovernanceReport.created_at).label("date"),
            func.count(GovernanceReport.id).label("count")
        )
        .filter(GovernanceReport.tenant_id == tenant_id)
        .group_by(func.date(GovernanceReport.created_at))
        .order_by(func.date(GovernanceReport.created_at))
        .all()
    )
    processing_trend = [
        {"date": str(row.date), "count": row.count}
        for row in trend_rows
    ]
    
    return {
        "reports_by_status": reports_by_status,
        "escalations_by_severity": escalations_by_severity,
        "raid_distribution": raid_distribution,
        "processing_trend": processing_trend
    }


# Helper Logic Functions for Phase 6 Governance Intelligence

def get_maturity_logic(db: Session, tenant_id: int):
    # 1. Policy ownership: percentage of RaidItem with suggested_owner_role set
    total_raid = db.query(RaidItem).filter(RaidItem.tenant_id == tenant_id).count()
    owned_raid = db.query(RaidItem).filter(RaidItem.tenant_id == tenant_id, RaidItem.suggested_owner_role != None, RaidItem.suggested_owner_role != "").count()
    policy_ownership = max(10, min(100, int(owned_raid / total_raid * 100))) if total_raid > 0 else 85
    
    # 2. Mitigation completion: completed/verified tasks / total tasks
    total_mit = db.query(MitigationTask).filter(MitigationTask.tenant_id == tenant_id).count()
    comp_mit = db.query(MitigationTask).filter(MitigationTask.tenant_id == tenant_id, MitigationTask.status.in_(["COMPLETED", "VERIFIED"])).count()
    mitigation_completion = max(10, min(100, int(comp_mit / total_mit * 100))) if total_mit > 0 else 75
    
    # 3. SLA compliance: tasks not overdue / total tasks
    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    overdue_mit = 0
    active_tasks = db.query(MitigationTask).filter(
        MitigationTask.tenant_id == tenant_id,
        MitigationTask.status.notin_(["COMPLETED", "VERIFIED"])
    ).all()
    for task in active_tasks:
        if task.target_date and task.target_date < today_str:
            overdue_mit += 1
    sla_compliance = max(10, min(100, int((total_mit - overdue_mit) / total_mit * 100))) if total_mit > 0 else 91
    
    # 4. Escalation closure: resolved/closed escalations / total escalations
    total_esc = db.query(EscalationItem).filter(EscalationItem.tenant_id == tenant_id).count()
    closed_esc = db.query(EscalationItem).filter(EscalationItem.tenant_id == tenant_id, EscalationItem.status.in_(["RESOLVED", "CLOSED"])).count()
    escalation_closure = max(10, min(100, int(closed_esc / total_esc * 100))) if total_esc > 0 else 69
    
    # 5. Risk reduction: (original_risk - current_risk) / original_risk
    total_original_risk = db.query(func.sum(RaidItem.risk_score)).filter(
        RaidItem.tenant_id == tenant_id,
        RaidItem.type.in_(["risk", "issue"])
    ).scalar() or 0
    total_current_risk = db.query(func.sum(RaidItem.current_risk_score)).filter(
        RaidItem.tenant_id == tenant_id,
        RaidItem.type.in_(["risk", "issue"])
    ).scalar() or 0
    if total_original_risk > 0:
        risk_reduction = max(10, min(100, int(round(((total_original_risk - total_current_risk) / total_original_risk) * 100))))
    else:
        risk_reduction = 73
        
    score = int((policy_ownership + mitigation_completion + sla_compliance + escalation_closure + risk_reduction) / 5)
    
    # Maturity tier
    if score >= 90:
        tier = "Optimized"
    elif score >= 75:
        tier = "Managed"
    elif score >= 55:
        tier = "Defined"
    elif score >= 35:
        tier = "Repeatable"
    else:
        tier = "Initial"
        
    # Benchmarking
    industry_average = 65
    delta = score - industry_average
    peer_percentile = max(1, min(99, int(50 + delta * 1.5)))
    
    # Appetite Alignment
    open_critical_risks = db.query(RaidItem).filter(
        RaidItem.tenant_id == tenant_id,
        RaidItem.type.in_(["risk", "issue"]),
        RaidItem.severity == "critical",
        RaidItem.current_risk_score > 0
    ).count()
    open_escalations = db.query(EscalationItem).filter(
        EscalationItem.tenant_id == tenant_id,
        EscalationItem.status.notin_(["RESOLVED", "CLOSED"])
    ).count()
    
    residual_ratio = (total_current_risk / total_original_risk) if total_original_risk > 0 else 0.3
    if open_critical_risks == 0 and open_escalations == 0 and risk_reduction >= 75:
        appetite_alignment = "Conservative"
    elif open_critical_risks <= 2 and open_escalations <= 2 and residual_ratio <= 0.5:
        appetite_alignment = "Balanced"
    else:
        appetite_alignment = "Aggressive"
        
    return {
        "score": score,
        "tier": tier,
        "dimensions": {
            "policy_ownership": policy_ownership,
            "mitigation_completion": mitigation_completion,
            "sla_compliance": sla_compliance,
            "escalation_closure": escalation_closure,
            "risk_reduction": risk_reduction
        },
        "benchmark": {
            "industry_average": industry_average,
            "peer_percentile": peer_percentile
        },
        "appetite_alignment": appetite_alignment
    }

def get_health_explanations_logic(db: Session, tenant_id: int):
    health_score = 100
    
    open_risks_count = db.query(RaidItem).filter(
        RaidItem.tenant_id == tenant_id,
        RaidItem.type.in_(["risk", "issue"]),
        RaidItem.severity.in_(["critical", "high"]),
        RaidItem.current_risk_score > 0
    ).count()
    health_score -= open_risks_count * 5
    
    open_escalations = db.query(EscalationItem).filter(
        EscalationItem.tenant_id == tenant_id,
        EscalationItem.status.notin_(["RESOLVED", "CLOSED"])
    ).count()
    health_score -= open_escalations * 8
    
    overdue_mitigations_count = 0
    active_tasks = db.query(MitigationTask).filter(
        MitigationTask.tenant_id == tenant_id,
        MitigationTask.status.notin_(["COMPLETED", "VERIFIED"])
    ).all()
    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    for task in active_tasks:
        if task.target_date and task.target_date < today_str:
            overdue_mitigations_count += 1
            
    health_score -= overdue_mitigations_count * 4
    
    verified_count = db.query(MitigationTask).filter(
        MitigationTask.tenant_id == tenant_id,
        MitigationTask.status == "VERIFIED"
    ).count()
    bonus = min(verified_count * 2, 15)
    health_score += bonus
    
    health_score = max(0, min(100, health_score))
    
    main_drivers = []
    if overdue_mitigations_count > 0:
        main_drivers.append({"description": f"{overdue_mitigations_count} overdue mitigation(s)", "impact": overdue_mitigations_count * 4})
    if open_escalations > 0:
        main_drivers.append({"description": f"{open_escalations} active escalation(s)", "impact": open_escalations * 8})
    if open_risks_count > 0:
        main_drivers.append({"description": f"{open_risks_count} unresolved critical/high risk(s)", "impact": open_risks_count * 5})
        
    positive_contributions = []
    if verified_count > 0:
        positive_contributions.append({"description": f"{verified_count} verified mitigation(s)", "impact": bonus})
        
    return {
        "health_score": health_score,
        "main_drivers": main_drivers,
        "positive_contributions": positive_contributions
    }

def get_executive_priorities_logic(db: Session, tenant_id: int):
    open_risks_count = db.query(RaidItem).filter(
        RaidItem.tenant_id == tenant_id,
        RaidItem.type.in_(["risk", "issue"]),
        RaidItem.severity.in_(["critical", "high"]),
        RaidItem.current_risk_score > 0
    ).count()
    
    open_escalations = db.query(EscalationItem).filter(
        EscalationItem.tenant_id == tenant_id,
        EscalationItem.status.notin_(["RESOLVED", "CLOSED"])
    ).count()
    
    overdue_mitigations_count = 0
    active_tasks = db.query(MitigationTask).filter(
        MitigationTask.tenant_id == tenant_id,
        MitigationTask.status.notin_(["COMPLETED", "VERIFIED"])
    ).all()
    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    for task in active_tasks:
        if task.target_date and task.target_date < today_str:
            overdue_mitigations_count += 1
            
    pending_reviews = db.query(GovernanceReport).filter(
        GovernanceReport.tenant_id == tenant_id,
        GovernanceReport.status == "PENDING_MANAGER_REVIEW"
    ).count()
    
    priorities = []
    
    # 1. Resolve open escalations
    score_esc = open_escalations * 8
    priorities.append({
        "title": "Resolve open escalations",
        "severity": "CRITICAL" if open_escalations > 1 else ("HIGH" if open_escalations > 0 else "LOW"),
        "count": open_escalations,
        "impact": "Potential SLA breach exposure",
        "priority_score": score_esc,
        "reason": f"Disputes outstanding: {open_escalations} open escalation item(s) awaiting resolution."
    })
    
    # 2. Review overdue mitigations
    score_overdue = overdue_mitigations_count * 9
    priorities.append({
        "title": "Review overdue mitigations",
        "severity": "HIGH" if overdue_mitigations_count > 0 else "LOW",
        "count": overdue_mitigations_count,
        "impact": "Prevents target SLA breaches and compliance exposure",
        "priority_score": score_overdue,
        "reason": f"SLA compliance impacted: {overdue_mitigations_count} overdue task(s) currently outstanding."
    })
    
    # 3. Resolve critical risks
    score_risks = open_risks_count * 11
    priorities.append({
        "title": "Resolve open critical/high risks",
        "severity": "CRITICAL" if open_risks_count > 0 else "LOW",
        "count": open_risks_count,
        "impact": "Protects critical business applications and assets",
        "priority_score": score_risks,
        "reason": f"Active threat exposure: {open_risks_count} unmitigated critical/high risk(s) in portfolio."
    })
    
    # 4. Approve pending reports
    score_reviews = pending_reviews * 4
    priorities.append({
        "title": "Approve pending reports",
        "severity": "MEDIUM" if pending_reviews > 0 else "LOW",
        "count": pending_reviews,
        "impact": "Completes draft ingestion pipeline review cycle",
        "priority_score": score_reviews,
        "reason": f"Workflow pending: {pending_reviews} reports awaiting manager review and approval."
    })
    
    priorities.sort(key=lambda x: x["priority_score"], reverse=True)
    return priorities

def get_root_cause_analytics_logic(db: Session, tenant_id: int):
    categories = ["AI Governance", "Security", "Privacy", "Vendor Risk", "Compliance", "Operational", "Financial", "Other"]
    dist = {c: 0 for c in categories}
    scores = {c: [] for c in categories}
    
    raid_items = db.query(RaidItem).filter(RaidItem.tenant_id == tenant_id).all()
    for item in raid_items:
        desc = (item.description or "").lower()
        if any(k in desc for k in ["ai", "llm", "model", "training data", "gpt", "claude", "neural", "algorithm"]):
            cat = "AI Governance"
        elif any(k in desc for k in ["security", "unauthorized", "vulnerability", "breach", "access", "bucket", "leak", "firewall", "encryption", "auth", "token", "password", "credential"]):
            cat = "Security"
        elif any(k in desc for k in ["privacy", "gdpr", "ccpa", "pii", "personal data", "consent"]):
            cat = "Privacy"
        elif any(k in desc for k in ["vendor", "third-party", "supplier", "contractor", "outsource"]):
            cat = "Vendor Risk"
        elif any(k in desc for k in ["compliance", "regulation", "sox", "hipaa", "audit", "policy", "legal", "framework"]):
            cat = "Compliance"
        elif any(k in desc for k in ["operational", "downtime", "outage", "process", "workflow", "capacity", "latency", "system", "performance"]):
            cat = "Operational"
        elif any(k in desc for k in ["financial", "budget", "cost", "revenue", "spending", "price", "billing"]):
            cat = "Financial"
        else:
            cat = "Other"
            
        dist[cat] += 1
        scores[cat].append(item.risk_score)
        
    category_distribution = dist
    category_risk_scores = {c: int(sum(scores[c])/len(scores[c])) if len(scores[c]) > 0 else 0 for c in categories}
    
    failure_patterns = []
    if dist["AI Governance"] > 0:
        failure_patterns.append("Unregulated AI usage and licensing ambiguity in training datasets.")
    if dist["Security"] > 0:
        failure_patterns.append("Vulnerable cloud access policies and missing data-at-rest encryption.")
    if dist["Compliance"] > 0:
        failure_patterns.append("SLA monitoring gaps leading to recurring corporate policy deviations.")
    if dist["Vendor Risk"] > 0:
        failure_patterns.append("Third-party vendor data-sharing compliance gaps.")
        
    if not failure_patterns:
        failure_patterns = ["All systems operate within acceptable governance parameters."]
        
    return {
        "category_distribution": category_distribution,
        "category_risk_scores": category_risk_scores,
        "failure_patterns": failure_patterns
    }


# Endpoints Definitions for Phase 6 Governance Intelligence

@router.get("/governance/maturity", response_model=GovernanceMaturityResponse)
def get_governance_maturity(
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    """Returns organizational maturity tier and dimensions scoring."""
    return get_maturity_logic(db, tenant_id)


@router.get("/governance/health-explanations", response_model=HealthExplanationsResponse)
def get_health_explanations(
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    """Explains factors impacting the current Governance Health Score."""
    return get_health_explanations_logic(db, tenant_id)


@router.get("/governance/executive-priorities", response_model=List[ExecutivePriorityItem])
def get_executive_priorities(
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    """Returns top executive priorities ranked by weighted scoring."""
    return get_executive_priorities_logic(db, tenant_id)


@router.get("/governance/root-cause-analytics", response_model=RootCauseAnalyticsResponse)
def get_root_cause_analytics(
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    """Categorizes and scores portfolio risks across 8 domains."""
    return get_root_cause_analytics_logic(db, tenant_id)


@router.get("/governance/portfolio-recommendations", response_model=StrategicRecommendationsResponse)
async def get_portfolio_recommendations(
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    """Generates AI-driven strategic governance recommendations."""
    from backend.app.services.ai.ai_service import AIService
    
    maturity = get_maturity_logic(db, tenant_id)
    rca = get_root_cause_analytics_logic(db, tenant_id)
    priorities = get_executive_priorities_logic(db, tenant_id)
    health_ex = get_health_explanations_logic(db, tenant_id)
    
    prio_str = ", ".join([f"{p['title']} ({p['count']})" for p in priorities])
    
    ai_service = AIService()
    prompt = (
        f"Generate strategic recommendations for a governance dashboard.\n"
        f"Current Health: {health_ex['health_score']}\n"
        f"Maturity Level: {maturity['score']} ({maturity['tier']})\n"
        f"Key Priorities: {prio_str}\n"
        f"Risk Category Distribution: {rca['category_distribution']}\n"
        f"Your response must include sections with headers: 'Quick Wins', 'Medium-Term', and 'Strategic'.\n"
        f"Provide 2 bullet points under each section."
    )
    
    resp_text = await ai_service.generate_text_completion(
        prompt=prompt,
        system_instruction="You are a senior enterprise GRC consultant. Return recommendations formatted as bullet points."
    )
    
    quick_wins = []
    medium_term = []
    strategic = []
    
    current_section = None
    for line in resp_text.split("\n"):
        line_clean = line.strip().lower()
        if not line_clean:
            continue
        if "quick win" in line_clean:
            current_section = "quick"
            continue
        elif "medium-term" in line_clean or "medium term" in line_clean:
            current_section = "medium"
            continue
        elif "strategic" in line_clean:
            current_section = "strategic"
            continue
            
        if line.strip().startswith(("-", "*", "1.", "2.")):
            content = line.strip().lstrip("-*12. ").strip()
            if content:
                if current_section == "quick":
                    quick_wins.append(content)
                elif current_section == "medium":
                    medium_term.append(content)
                elif current_section == "strategic":
                    strategic.append(content)
                    
    open_risks_count = sum(rca["category_distribution"].values())
    if not quick_wins:
        quick_wins = [
            f"Assign owners to all {open_risks_count} unassigned RAID items in the catalog." if open_risks_count > 0 else "Audit low-priority risk warnings.",
            "Integrate Slack/Teams webhook alerts for critical SLA triggers."
        ]
    if not medium_term:
        medium_term = [
            "Establish automated SLA thresholds for mitigation owner notifications.",
            "Standardize tenant-specific data retention and archiving policies."
        ]
    if not strategic:
        strategic = [
            "Form a dedicated AI Governance Committee to review training model licensing.",
            "Establish continuous verification cycles to transition to the Optimized maturity tier."
        ]
        
    return {
        "quick_wins": quick_wins[:3],
        "medium_term": medium_term[:3],
        "strategic": strategic[:3]
    }


@router.get("/governance/trends", response_model=GovernanceTrendsResponse)
def get_governance_trends(
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    """Returns chronological snapshots of organizational governance."""
    snapshots = db.query(GovernanceTrendSnapshot).filter(
        GovernanceTrendSnapshot.tenant_id == tenant_id
    ).order_by(GovernanceTrendSnapshot.timestamp.asc()).all()
    
    points = []
    for s in snapshots:
        points.append({
            "date": s.timestamp.strftime("%Y-%m-%d"),
            "health_score": s.health_score,
            "maturity_score": s.maturity_score,
            "risk_exposure": s.risk_exposure,
            "mitigation_effectiveness_pct": s.mitigation_effectiveness_pct,
            "sla_breaches": s.sla_breaches,
            "open_escalations": s.open_escalations,
            "verified_mitigations": s.verified_mitigations,
            "critical_risks": s.critical_risks,
            "notification_volume": s.notification_volume
        })
        
    if not points:
        for day in range(30, 0, -1):
            date_str = (datetime.utcnow() - timedelta(days=day)).strftime("%Y-%m-%d")
            points.append({
                "date": date_str,
                "health_score": 75 + (day % 3) * 2,
                "maturity_score": 70 + (day % 2),
                "risk_exposure": 300 - day * 3,
                "mitigation_effectiveness_pct": 65.0 + (day % 5) * 2,
                "sla_breaches": max(0, 5 - day // 6),
                "open_escalations": max(0, 3 - day // 10),
                "verified_mitigations": day // 3,
                "critical_risks": max(0, 4 - day // 8),
                "notification_volume": 15 + (day % 4) * 3
            })
            
    return {"trend_points": points}


@router.get("/governance/executive-briefing", response_model=ExecutiveBriefingResponse)
def get_executive_briefing(
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    """Generates board-ready executive summaries and state briefs."""
    org = db.query(Organization).filter(Organization.id == tenant_id).first()
    tenant_name = org.name if org else "Enterprise Demo Organization"
    
    maturity = get_maturity_logic(db, tenant_id)
    health_ex = get_health_explanations_logic(db, tenant_id)
    rca = get_root_cause_analytics_logic(db, tenant_id)
    
    total_docs = db.query(Document).filter(Document.tenant_id == tenant_id).count()
    open_risks_count = sum(rca["category_distribution"].values())
    
    overdue_count = 0
    active_tasks = db.query(MitigationTask).filter(
        MitigationTask.tenant_id == tenant_id,
        MitigationTask.status.notin_(["COMPLETED", "VERIFIED"])
    ).all()
    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    for task in active_tasks:
        if task.target_date and task.target_date < today_str:
            overdue_count += 1
            
    open_esc = db.query(EscalationItem).filter(
        EscalationItem.tenant_id == tenant_id,
        EscalationItem.status.notin_(["RESOLVED", "CLOSED"])
    ).count()
    
    executive_summary = (
        f"Overall organizational governance stands at a Health Score of {health_ex['health_score']} (Strong) "
        f"and a Maturity Score of {maturity['score']} ({maturity['tier']}) under tenant '{tenant_name}'."
    )
    current_state = (
        f"Active tenant context contains {total_docs} documents processed, with a cumulative risk reduction of "
        f"{maturity['dimensions']['risk_reduction']}% achieved through verified mitigations."
    )
    key_risks = (
        f"Active portfolio exposure includes {open_risks_count} unresolved risks. Security vulnerabilities and "
        f"unassigned owners represent the main risk vector."
    )
    operational_concerns = (
        f"We currently have {overdue_count} overdue mitigation tasks causing SLA breaches, and {open_esc} open "
        f"corporate-level escalations."
    )
    recommendations = (
        "Establish an AI Governance review board, publish policies on LLM model training datasets, and patch "
        "cloud access permissions."
    )
    next_30_days = (
        "Remediate the cloud storage access controls, resolve outstanding escalations, and verify completed mitigations."
    )
    
    full_markdown = (
        f"# EXECUTIVE GOVERNANCE BRIEFING\n\n"
        f"**Tenant:** {tenant_name}  \n"
        f"**Generated:** {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}  \n"
        f"**Governance Health:** {health_ex['health_score']}/100  \n"
        f"**Maturity Level:** {maturity['score']} ({maturity['tier']})  \n"
        f"**Risk Appetite:** {maturity['appetite_alignment']}  \n\n"
        f"---\n\n"
        f"### 1. Executive Summary\n"
        f"{executive_summary}\n\n"
        f"### 2. Current State & Coverage\n"
        f"{current_state}\n\n"
        f"### 3. Key Risks & Exposure\n"
        f"{key_risks}\n\n"
        f"### 4. Operational Concerns & SLA Breaches\n"
        f"{operational_concerns}\n\n"
        f"### 5. Recommendations\n"
        f"{recommendations}\n\n"
        f"### 6. Next 30 Days Plan\n"
        f"{next_30_days}\n"
    )
    
    return {
        "executive_summary": executive_summary,
        "current_state": current_state,
        "key_risks": key_risks,
        "operational_concerns": operational_concerns,
        "recommendations": recommendations,
        "next_30_days": next_30_days,
        "full_markdown": full_markdown
    }


@router.post("/governance/copilot", response_model=CopilotResponse)
async def get_governance_copilot(
    payload: CopilotRequest,
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant),
    x_user_role: str = Depends(get_current_role)
):
    """Role-aware natural language assistant for governance intelligence."""
    from backend.app.services.ai.ai_service import AIService
    
    maturity = get_maturity_logic(db, tenant_id)
    health_ex = get_health_explanations_logic(db, tenant_id)
    rca = get_root_cause_analytics_logic(db, tenant_id)
    priorities = get_executive_priorities_logic(db, tenant_id)
    
    # Retrieve relevant document context via FAISS RAG
    from backend.app.services.rag.retrieval import RetrievalService
    rag_context = RetrievalService.retrieve_relevant_context(
        query=payload.query,
        document_id=None,
        top_k=3
    )
    
    prio_str = "\n".join([f"- {p['title']} (Count: {p['count']}, Severity: {p['severity']}, Impact: {p['impact']}, Score: {p['priority_score']})" for p in priorities])
    drivers_str = "\n".join([f"- {d['description']} (-{d['impact']} pts)" for d in health_ex["main_drivers"]])
    bonus_str = "\n".join([f"- {b['description']} (+{b['impact']} pts)" for b in health_ex["positive_contributions"]])
    
    system_instruction = (
        f"You are the Governance Intelligence Executive Copilot, an AI assistant built into the Enterprise Governance Intelligence Platform.\n"
        f"Your task is to answer user queries based on the organization's governance metadata and retrieved document context.\n"
        f"Strictly use the following context. Do not make up any other data.\n"
        f"User Role: {x_user_role}\n"
        f"Relevant Document Context (RAG):\n{rag_context}\n"
        f"Governance Health Score: {health_ex['health_score']}\n"
        f"Governance Maturity Score: {maturity['score']} ({maturity['tier']})\n"
        f"Maturity Dimensions:\n"
        f"  - Policy Ownership: {maturity['dimensions']['policy_ownership']}\n"
        f"  - Mitigation Completion: {maturity['dimensions']['mitigation_completion']}\n"
        f"  - SLA Compliance: {maturity['dimensions']['sla_compliance']}\n"
        f"  - Escalation Closure: {maturity['dimensions']['escalation_closure']}\n"
        f"  - Risk Reduction: {maturity['dimensions']['risk_reduction']}\n"
        f"Appetite Alignment: {maturity['appetite_alignment']}\n"
        f"Active Priorities:\n{prio_str}\n"
        f"Health Score Negative Drivers:\n{drivers_str}\n"
        f"Health Score Positive Contributions:\n{bonus_str}\n"
        f"Root Cause Category Distribution: {rca['category_distribution']}\n"
        f"Root Cause Category Risk Scores: {rca['category_risk_scores']}\n"
        f"Tailor your response to the user's role ({x_user_role}).\n"
        f"If the user is an Analyst, focus on specific task completions, overdue mitigations, and operational details.\n"
        f"If the user is a Manager, focus on team queue reviews, pending report reviews, and escalation routing.\n"
        f"If the user is a Governance Lead or Executive, focus on high-level risk exposure, board pack preparation, strategic alignment, and peer benchmarking."
    )
    
    ai_service = AIService()
    response_text = await ai_service.generate_text_completion(
        prompt=payload.query,
        system_instruction=system_instruction
    )
    
    return {"response": response_text}


@router.get("/governance/escalations", response_model=List[EscalationItemResponse])
def list_escalations(
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    """Lists escalations, optionally filtered by status (open, routed)."""
    query = db.query(EscalationItem).options(
        joinedload(EscalationItem.report).joinedload(GovernanceReport.document)
    ).filter(EscalationItem.tenant_id == tenant_id)
    if status:
        query = query.filter(EscalationItem.status == status)
    return query.order_by(EscalationItem.created_at.desc()).all()

@router.post("/governance/escalations/{id}/route", response_model=EscalationItemResponse)
def route_escalation(
    id: int,
    payload: EscalationRouteRequest,
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role)
):
    """Routes an escalation item to a specified target stakeholder."""
    if x_user_role not in ("Manager", "Governance Lead"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Managers or Governance Leads can route escalations."
        )
    esc = db.query(EscalationItem).options(
        joinedload(EscalationItem.report).joinedload(GovernanceReport.document)
    ).filter(EscalationItem.id == id).first()
    if not esc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Escalation item not found")
        
    esc.status = "ASSIGNED"
    esc.routing_target = payload.routing_target
    
    # Add Audit log
    report = db.query(GovernanceReport).filter(GovernanceReport.id == esc.report_id).first()
    doc_id = report.document_id if report else None
    
    audit = AuditLog(
        document_id=doc_id,
        governance_report_id=esc.report_id,
        event="Escalation Routed",
        user="reviewer_user",
        user_role=x_user_role,
        action="escalation routing",
        entity_type="escalation",
        entity_id=esc.id,
        details=f"Escalation routed to {payload.routing_target}. Issue: {esc.description[:100]}"
    )
    
    db.add(audit)
    
    # Create notification for routed target
    notif_target = payload.routing_target.title() if payload.routing_target else "Governance Lead"
    if "Lead" in notif_target:
        notif_target = "Governance Lead"
    notif = Notification(
        severity="MEDIUM",
        notification_type="ESCALATION_ASSIGNED",
        title="Escalation Routed",
        message=f"Escalation item {esc.id} has been routed to you. Description: {esc.description[:100]}.",
        recipient_role=notif_target,
        related_entity_type="escalation",
        related_entity_id=esc.id,
        read_status=False
    )
    db.add(notif)
    
    db.commit()
    db.refresh(esc)
    
    logger.info(f"Escalation ID {esc.id} successfully routed to {payload.routing_target}")
    return esc

@router.patch("/governance/reports/{id}/assign", response_model=GovernanceReportResponse)
def assign_governance_report(
    id: int,
    payload: ReportAssignRequest,
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role)
):
    """Assigns a governance report to a specific user/role."""
    if x_user_role != "Manager":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Managers can assign reports."
        )
    report = db.query(GovernanceReport).filter(GovernanceReport.id == id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Governance report not found")
        
    report.assigned_to = payload.assigned_to
    
    audit = AuditLog(
        document_id=report.document_id,
        governance_report_id=report.id,
        event="Report Assigned",
        user="manager_user",
        user_role=x_user_role,
        action="report assignment",
        entity_type="report",
        entity_id=report.id,
        details=f"Report assigned to {payload.assigned_to}."
    )
    db.add(audit)

    # Create notification for assignee
    notif_role = payload.assigned_to.title() if payload.assigned_to else "Analyst"
    if "Lead" in notif_role:
        notif_role = "Governance Lead"
    notif = Notification(
        severity="LOW",
        notification_type="GOVERNANCE_ALERT",
        title="Governance Report Assigned",
        message=f"Governance Report V{report.version} has been assigned to you.",
        recipient_role=notif_role,
        related_entity_type="report",
        related_entity_id=report.id,
        read_status=False
    )
    db.add(notif)

    db.commit()
    db.refresh(report)
    return report

@router.patch("/governance/reports/{id}/escalate", response_model=GovernanceReportResponse)
def escalate_governance_report(
    id: int,
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role)
):
    """Escalates a report, transitioning its owner and status, and assigning any associated escalations."""
    if x_user_role != "Manager":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Managers can escalate reports."
        )
    report = db.query(GovernanceReport).filter(GovernanceReport.id == id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Governance report not found")
        
    report.status = "ESCALATED"
    report.assigned_to = "Governance Lead"
    
    # Automatically assign associated escalation items to Governance Lead
    escalations = db.query(EscalationItem).filter(EscalationItem.report_id == report.id).all()
    for esc in escalations:
        esc.status = "ASSIGNED"
        esc.assigned_to = "Governance Lead"
        esc.raised_by = "Manager"
        
    audit = AuditLog(
        document_id=report.document_id,
        governance_report_id=report.id,
        event="Report Escalated",
        user="manager_user",
        user_role=x_user_role,
        action="report escalation",
        entity_type="report",
        entity_id=report.id,
        details="Report escalated to Governance Lead. Associated escalation items assigned."
    )
    db.add(audit)

    # Create notification for Governance Lead
    notif = Notification(
        severity="HIGH",
        notification_type="ESCALATION_ASSIGNED",
        title="Governance Report Escalated",
        message=f"Governance Report V{report.version} has been escalated to you.",
        recipient_role="Governance Lead",
        related_entity_type="report",
        related_entity_id=report.id,
        read_status=False
    )
    db.add(notif)

    db.commit()
    db.refresh(report)
    return report

@router.patch("/governance/escalations/{id}/assign", response_model=EscalationItemResponse)
def assign_escalation(
    id: int,
    payload: EscalationAssignRequest,
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role)
):
    """Assigns an escalation item to a specific user/role."""
    if x_user_role != "Governance Lead":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Governance Leads can assign escalations."
        )
    esc = db.query(EscalationItem).filter(EscalationItem.id == id).first()
    if not esc:
        raise HTTPException(status_code=404, detail="Escalation item not found")
        
    esc.assigned_to = payload.assigned_to
    esc.status = "ASSIGNED"
    
    audit = AuditLog(
        governance_report_id=esc.report_id,
        event="Escalation Assigned",
        user="gov_lead_user",
        user_role=x_user_role,
        action="escalation assignment",
        entity_type="escalation",
        entity_id=esc.id,
        details=f"Escalation assigned to {payload.assigned_to}."
    )
    db.add(audit)

    # Create notification for assignee
    notif_role = payload.assigned_to.title() if payload.assigned_to else "Governance Lead"
    if "Lead" in notif_role:
        notif_role = "Governance Lead"
    notif = Notification(
        severity="MEDIUM",
        notification_type="ESCALATION_ASSIGNED",
        title="Escalation Assigned",
        message=f"Escalation item {esc.id} has been assigned to you.",
        recipient_role=notif_role,
        related_entity_type="escalation",
        related_entity_id=esc.id,
        read_status=False
    )
    db.add(notif)

    db.commit()
    db.refresh(esc)
    return esc

@router.patch("/governance/escalations/{id}/resolve", response_model=EscalationItemResponse)
def resolve_escalation(
    id: int,
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role)
):
    """Resolves an active escalation item."""
    if x_user_role != "Governance Lead":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Governance Leads can resolve escalations."
        )
    esc = db.query(EscalationItem).filter(EscalationItem.id == id).first()
    if not esc:
        raise HTTPException(status_code=404, detail="Escalation item not found")
        
    esc.status = "RESOLVED"
    esc.resolved_by = "Governance Lead"
    
    audit = AuditLog(
        governance_report_id=esc.report_id,
        event="Escalation Resolved",
        user="gov_lead_user",
        user_role=x_user_role,
        action="escalation resolution",
        entity_type="escalation",
        entity_id=esc.id,
        details="Escalation resolved by Governance Lead."
    )
    db.add(audit)

    # Create notification for Manager
    notif = Notification(
        severity="LOW",
        notification_type="GOVERNANCE_ALERT",
        title="Escalation Resolved",
        message=f"Escalation item {esc.id} has been resolved by Governance Lead.",
        recipient_role="Manager",
        related_entity_type="escalation",
        related_entity_id=esc.id,
        read_status=False
    )
    db.add(notif)

    db.commit()
    db.refresh(esc)
    return esc

@router.patch("/governance/escalations/{id}/close", response_model=EscalationItemResponse)
def close_escalation(
    id: int,
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role)
):
    """Closes and archives an escalation item."""
    if x_user_role != "Governance Lead":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Governance Leads can close escalations."
        )
    esc = db.query(EscalationItem).filter(EscalationItem.id == id).first()
    if not esc:
        raise HTTPException(status_code=404, detail="Escalation item not found")
        
    esc.status = "CLOSED"
    
    audit = AuditLog(
        governance_report_id=esc.report_id,
        event="Escalation Closed",
        user="gov_lead_user",
        user_role=x_user_role,
        action="escalation closure",
        entity_type="escalation",
        entity_id=esc.id,
        details="Escalation formally closed."
    )
    db.add(audit)
    db.commit()
    db.refresh(esc)
    return esc

@router.get("/governance/audit-events", response_model=List[AuditLogResponse])
def list_audit_events(db: Session = Depends(get_db)):
    """Lists all workflow audit logs (timeline items) chronologically."""
    return db.query(AuditLog).order_by(AuditLog.timestamp.desc()).all()


def calculate_sla_status(status: str, target_date: Optional[str]) -> str:
    if status in ("COMPLETED", "VERIFIED"):
        return "ON_TRACK"
    if not target_date:
        return "ON_TRACK"
    try:
        due_date = datetime.strptime(target_date, "%Y-%m-%d").date()
        today = datetime.utcnow().date()
        if due_date < today:
            return "OVERDUE"
        elif (due_date - today).days <= 3:
            return "AT_RISK"
    except ValueError:
        pass
    return "ON_TRACK"


def recalculate_raid_item_risk(db: Session, raid_item_id: int):
    raid_item = db.query(RaidItem).filter(RaidItem.id == raid_item_id).first()
    if not raid_item:
        return
    
    # Get all verified mitigation tasks linked to this raid item
    verified_tasks = db.query(MitigationTask).filter(
        MitigationTask.related_raid_item_id == raid_item_id,
        MitigationTask.status == "VERIFIED"
    ).all()
    
    total_effectiveness = sum(task.effectiveness for task in verified_tasks)
    # Cap total effectiveness at 80% (0.80) to guarantee a 20% residual risk floor
    if total_effectiveness > 80:
        total_effectiveness = 80
        
    residual_mult = 1.0 - (total_effectiveness / 100.0)
    
    # Calculate current risk score
    new_risk_score = int(round(raid_item.risk_score * residual_mult))
    raid_item.current_risk_score = max(0, new_risk_score)
    db.commit()


def log_mitigation_audit(db: Session, task: MitigationTask, event: str, user_role: str, user: str, details: str, action: str):
    # Retrieve report_id and document_id from the related RAID item
    raid_item = db.query(RaidItem).filter(RaidItem.id == task.related_raid_item_id).first()
    doc_id = None
    report_id = None
    if raid_item and raid_item.report:
        report_id = raid_item.report.id
        doc_id = raid_item.report.document_id
        
    audit = AuditLog(
        document_id=doc_id,
        governance_report_id=report_id,
        event=event,
        user=user,
        user_role=user_role,
        action=action,
        entity_type="mitigation",
        entity_id=task.id,
        details=details
    )
    db.add(audit)
    db.commit()


@router.get("/mitigations", response_model=List[MitigationTaskResponse])
def list_mitigations(
    status: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    owner_role: Optional[str] = Query(None),
    related_raid_item_id: Optional[int] = Query(None),
    related_escalation_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    query = db.query(MitigationTask).filter(MitigationTask.tenant_id == tenant_id)
    if status:
        query = query.filter(MitigationTask.status == status)
    if priority:
        query = query.filter(MitigationTask.priority == priority)
    if owner_role:
        query = query.filter(MitigationTask.owner_role == owner_role)
    if related_raid_item_id:
        query = query.filter(MitigationTask.related_raid_item_id == related_raid_item_id)
    if related_escalation_id:
        query = query.filter(MitigationTask.related_escalation_id == related_escalation_id)
    
    tasks = query.order_by(MitigationTask.created_at.desc()).all()
    
    # Update SLA status dynamically
    for task in tasks:
        task.sla_status = calculate_sla_status(task.status, task.target_date)
        
    return tasks


@router.get("/mitigations/{id}", response_model=MitigationTaskResponse)
def get_mitigation(id: int, db: Session = Depends(get_db)):
    task = db.query(MitigationTask).filter(MitigationTask.id == id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Mitigation task not found")
    task.sla_status = calculate_sla_status(task.status, task.target_date)
    return task


@router.put("/mitigations/{id}", response_model=MitigationTaskResponse)
def update_mitigation(
    id: int,
    payload: MitigationTaskUpdateRequest,
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role)
):
    task = db.query(MitigationTask).filter(MitigationTask.id == id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Mitigation task not found")
        
    role = x_user_role
    
    # 1. Check if task is currently verified (only Governance Lead can modify it)
    if task.status == "VERIFIED" and role != "Governance Lead":
        raise HTTPException(
            status_code=403,
            detail="Only Governance Lead can modify a verified mitigation task."
        )
        
    # 2. Check if task is currently completed (only Manager or Governance Lead can modify it)
    if task.status == "COMPLETED" and role == "Analyst":
        raise HTTPException(
            status_code=403,
            detail="Analysts cannot modify completed mitigation tasks."
        )
        
    # Check target status if it changes
    target_status = payload.status
    if target_status and target_status != task.status:
        # 3. Only Governance Lead can mark VERIFIED
        if target_status == "VERIFIED" and role != "Governance Lead":
            raise HTTPException(
                status_code=403,
                detail="Only Governance Lead can mark tasks as verified."
            )
        # 4. Analysts cannot mark COMPLETED
        if target_status == "COMPLETED" and role == "Analyst":
            raise HTTPException(
                status_code=403,
                detail="Analysts cannot mark tasks as completed."
            )
            
    # Save old status to see if it changes to/from VERIFIED
    old_status = task.status
    
    # Update fields if provided
    if payload.title is not None:
        task.title = payload.title
    if payload.description is not None:
        task.description = payload.description
        
    # Log assignments if owner details change
    owner_changed = False
    if payload.owner_role is not None and payload.owner_role != task.owner_role:
        task.owner_role = payload.owner_role
        owner_changed = True
    if payload.owner_name is not None and payload.owner_name != task.owner_name:
        task.owner_name = payload.owner_name
        owner_changed = True
    if payload.target_date is not None and payload.target_date != task.target_date:
        task.target_date = payload.target_date
        owner_changed = True
        
    if owner_changed:
        log_mitigation_audit(
            db, task, "MITIGATION_ASSIGNED", role, f"{role.lower()}_user",
            f"Mitigation task owner or schedule updated. Role: {task.owner_role}, Owner: {task.owner_name or 'Unassigned'}, Due: {task.target_date or 'None'}",
            "mitigation assignment"
        )
        notif = Notification(
            severity="LOW" if task.priority in ("P3", "P4") else "MEDIUM",
            notification_type="MITIGATION_ASSIGNED",
            title="Mitigation Task Assigned",
            message=f"Mitigation task '{task.title}' has been assigned to you.",
            recipient_role=task.owner_role,
            related_entity_type="mitigation",
            related_entity_id=task.id,
            read_status=False
        )
        db.add(notif)
        
    if payload.priority is not None:
        task.priority = payload.priority
    if payload.completion_percentage is not None:
        task.completion_percentage = payload.completion_percentage
    if payload.effectiveness is not None:
        task.effectiveness = payload.effectiveness
        
    if target_status is not None and target_status != old_status:
        task.status = target_status
        
        # Handle date tracking and notifications
        if target_status == "COMPLETED":
            task.completed_at = datetime.utcnow()
            notif = Notification(
                severity="MEDIUM",
                notification_type="GOVERNANCE_ALERT",
                title="Mitigation Task Completed",
                message=f"Mitigation task '{task.title}' has been completed by {role} and is ready for verification.",
                recipient_role="Governance Lead",
                related_entity_type="mitigation",
                related_entity_id=task.id,
                read_status=False
            )
            db.add(notif)
        elif target_status == "VERIFIED":
            task.verified_at = datetime.utcnow()
            if not task.completed_at:
                task.completed_at = datetime.utcnow()
            notif = Notification(
                severity="LOW",
                notification_type="GOVERNANCE_ALERT",
                title="Mitigation Task Verified",
                message=f"Mitigation task '{task.title}' has been verified by Governance Lead.",
                recipient_role=task.owner_role,
                related_entity_type="mitigation",
                related_entity_id=task.id,
                read_status=False
            )
            db.add(notif)
        elif old_status in ("COMPLETED", "VERIFIED") and target_status in ("IN_PROGRESS", "PLANNED", "BLOCKED"):
            task.completed_at = None
            task.verified_at = None
            notif = Notification(
                severity="MEDIUM",
                notification_type="MITIGATION_ASSIGNED",
                title="Mitigation Task Reopened",
                message=f"Mitigation task '{task.title}' has been reopened and set to {target_status.lower()}.",
                recipient_role=task.owner_role,
                related_entity_type="mitigation",
                related_entity_id=task.id,
                read_status=False
            )
            db.add(notif)
        else:
            task.completed_at = None
            task.verified_at = None
            
        # Log audit timeline event
        event_map = {
            "IN_PROGRESS": ("MITIGATION_STARTED", "mitigation started", "Mitigation task set to in progress."),
            "BLOCKED": ("MITIGATION_BLOCKED", "mitigation blocked", "Mitigation task set to blocked."),
            "COMPLETED": ("MITIGATION_COMPLETED", "mitigation completion", "Mitigation task marked completed."),
            "VERIFIED": ("MITIGATION_VERIFIED", "mitigation verification", "Mitigation task verified.")
        }
        if target_status in event_map:
            evt, act, det = event_map[target_status]
            log_mitigation_audit(db, task, evt, role, f"{role.lower()}_user", det, act)
        elif old_status in ("COMPLETED", "VERIFIED"):
            log_mitigation_audit(db, task, "MITIGATION_REOPENED", role, f"{role.lower()}_user", "Mitigation task reopened.", "mitigation reopening")
            
    # Recompute SLA status
    task.sla_status = calculate_sla_status(task.status, task.target_date)
    task.updated_at = datetime.utcnow()
    db.commit()
    
    # Trigger risk recalculation if verified status changed
    if (old_status == "VERIFIED" and task.status != "VERIFIED") or (old_status != "VERIFIED" and task.status == "VERIFIED"):
        recalculate_raid_item_risk(db, task.related_raid_item_id)
        
    db.refresh(task)
    return task


@router.post("/mitigations/{id}/verify", response_model=MitigationTaskResponse)
def verify_mitigation(
    id: int,
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role)
):
    if x_user_role != "Governance Lead":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Governance Lead can verify mitigation tasks."
        )
    task = db.query(MitigationTask).filter(MitigationTask.id == id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Mitigation task not found")
        
    if task.status != "VERIFIED":
        task.status = "VERIFIED"
        task.verified_at = datetime.utcnow()
        if not task.completed_at:
            task.completed_at = datetime.utcnow()
        task.sla_status = "ON_TRACK"
        task.updated_at = datetime.utcnow()
        
        # Create notification for task owner
        notif = Notification(
            severity="LOW",
            notification_type="GOVERNANCE_ALERT",
            title="Mitigation Task Verified",
            message=f"Mitigation task '{task.title}' has been verified by Governance Lead.",
            recipient_role=task.owner_role,
            related_entity_type="mitigation",
            related_entity_id=task.id,
            read_status=False
        )
        db.add(notif)
        
        db.commit()
        
        log_mitigation_audit(
            db, task, "MITIGATION_VERIFIED", x_user_role, "gov_lead_user",
            "Mitigation task verified and approved by Governance Lead.", "mitigation verification"
        )
        
        recalculate_raid_item_risk(db, task.related_raid_item_id)
        db.refresh(task)
        
    return task


@router.post("/mitigations/{id}/reopen", response_model=MitigationTaskResponse)
def reopen_mitigation(
    id: int,
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role)
):
    if x_user_role != "Governance Lead":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Governance Lead can reopen mitigation tasks."
        )
    task = db.query(MitigationTask).filter(MitigationTask.id == id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Mitigation task not found")
        
    task.status = "IN_PROGRESS"
    task.completed_at = None
    task.verified_at = None
    task.sla_status = calculate_sla_status(task.status, task.target_date)
    task.updated_at = datetime.utcnow()
    
    # Create notification for task owner
    notif = Notification(
        severity="MEDIUM",
        notification_type="MITIGATION_ASSIGNED",
        title="Mitigation Task Reopened",
        message=f"Mitigation task '{task.title}' has been reopened by Governance Lead.",
        recipient_role=task.owner_role,
        related_entity_type="mitigation",
        related_entity_id=task.id,
        read_status=False
    )
    db.add(notif)
    
    db.commit()
    
    log_mitigation_audit(
        db, task, "MITIGATION_REOPENED", x_user_role, "gov_lead_user",
        "Mitigation task reopened and marked in progress by Governance Lead.", "mitigation reopening"
    )
    
    recalculate_raid_item_risk(db, task.related_raid_item_id)
    db.refresh(task)
    
    return task


@router.get("/notifications", response_model=List[NotificationResponse])
def list_notifications(
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role),
    tenant_id: int = Depends(get_current_tenant)
):
    """Lists notifications for the active user role, syncing dynamic SLA alerts first."""
    check_dynamic_sla_notifications(db)
    role = x_user_role
    return db.query(Notification).filter(
        Notification.recipient_role == role,
        Notification.tenant_id == tenant_id
    ).order_by(Notification.created_at.desc()).all()


@router.put("/notifications/{id}/read", response_model=NotificationResponse)
def read_notification(
    id: int,
    payload: NotificationReadRequest,
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role)
):
    """Marks a single notification as read/unread if assigned to active user role."""
    notif = db.query(Notification).filter(
        Notification.id == id,
        Notification.recipient_role == x_user_role
    ).first()
    if not notif:
        raise HTTPException(status_code=404, detail="Notification not found")
    notif.read_status = payload.read_status
    db.commit()
    db.refresh(notif)
    return notif


@router.put("/notifications/read-all")
def read_all_notifications(
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role)
):
    """Marks all unread notifications for the active user role as read."""
    db.query(Notification).filter(
        Notification.recipient_role == x_user_role,
        Notification.read_status == False
    ).update({Notification.read_status: True}, synchronize_session=False)
    db.commit()
    return {"message": "All notifications marked as read."}


@router.get("/inbox", response_model=InboxResponse)
def get_inbox(
    db: Session = Depends(get_db),
    x_user_role: str = Depends(get_current_role),
    tenant_id: int = Depends(get_current_tenant)
):
    """Aggregates pending operational workflow actions and assignments scoped to active role."""
    # Sync dynamic SLA alerts first
    check_dynamic_sla_notifications(db)
    role = x_user_role
    
    # 1. Pending Reviews:
    if role == "Manager":
        pending_reviews = db.query(GovernanceReport).filter(
            GovernanceReport.status == "PENDING_MANAGER_REVIEW",
            GovernanceReport.is_latest == True,
            GovernanceReport.tenant_id == tenant_id
        ).all()
    elif role == "Governance Lead":
        pending_reviews = db.query(GovernanceReport).filter(
            GovernanceReport.status == "ESCALATED",
            GovernanceReport.is_latest == True,
            GovernanceReport.tenant_id == tenant_id
        ).all()
    else:  # Analyst
        pending_reviews = db.query(GovernanceReport).filter(
            GovernanceReport.review_status == "changes_requested",
            GovernanceReport.is_latest == True,
            GovernanceReport.tenant_id == tenant_id
        ).all()
        
    # 2. Assigned Escalations:
    assigned_escalations = db.query(EscalationItem).filter(
        EscalationItem.assigned_to == role,
        EscalationItem.status.notin_(["RESOLVED", "CLOSED"]),
        EscalationItem.tenant_id == tenant_id
    ).all()
    
    # 3. Assigned Mitigations:
    assigned_mitigations = db.query(MitigationTask).filter(
        MitigationTask.owner_role == role,
        MitigationTask.status.in_(["PLANNED", "IN_PROGRESS", "BLOCKED"]),
        MitigationTask.tenant_id == tenant_id
    ).all()
    
    # Update SLA status dynamically
    for task in assigned_mitigations:
        task.sla_status = calculate_sla_status(task.status, task.target_date)
        
    # 4. Pending Verifications (Governance Lead only):
    if role == "Governance Lead":
        pending_verifications = db.query(MitigationTask).filter(
            MitigationTask.status == "COMPLETED",
            MitigationTask.tenant_id == tenant_id
        ).all()
        for task in pending_verifications:
            task.sla_status = calculate_sla_status(task.status, task.target_date)
    else:
        pending_verifications = []
        
    return {
        "pending_reviews": pending_reviews,
        "assigned_escalations": assigned_escalations,
        "assigned_mitigations": assigned_mitigations,
        "pending_verifications": pending_verifications
    }


@router.post("/demo-data/generate", status_code=status.HTTP_200_OK)
def generate_demo_data(
    payload: DemoDataGenerateRequest,
    db: Session = Depends(get_db)
):
    """Clean-wipes the database tables and generates realistic small, medium, or enterprise datasets."""
    size = payload.size
    
    # 1. Truncate tables in dependency order
    try:
        db.query(Notification).delete()
        db.query(AuditLog).delete()
        db.query(MitigationTask).delete()
        db.query(MeetingAction).delete()
        db.query(EscalationItem).delete()
        db.query(RaidItem).delete()
        db.query(GovernanceReport).delete()
        db.query(WorkflowJob).delete()
        db.query(Document).delete()
        db.query(GovernanceTrendSnapshot).delete()
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to clean database: {e}")
        raise HTTPException(status_code=500, detail="Database cleanup failed during seeding.")
        
    # 2. Seed data based on size
    from datetime import timedelta
    import json
    
    # Let's define document templates
    doc_templates = [
        {"filename": "vendor_security_audit_v2.pdf", "type": "pdf"},
        {"filename": "FY26_Q1_Steering_Committee_Pack.docx", "type": "docx"},
        {"filename": "project_aurora_risk_register.txt", "type": "txt"},
        {"filename": "cloud_migration_governance_memo.pdf", "type": "pdf"},
        {"filename": "SLA_breach_incident_response.pdf", "type": "pdf"},
        {"filename": "executive_escalation_memo_sec.docx", "type": "docx"},
        {"filename": "data_leakage_prevent_plan.pdf", "type": "pdf"},
        {"filename": "annual_it_compliance_assessment.txt", "type": "txt"},
        {"filename": "external_api_dependency_mapping.pdf", "type": "pdf"},
        {"filename": "incident_rca_db_outage.pdf", "type": "pdf"},
        {"filename": "steering_committee_minutes_may26.pdf", "type": "pdf"},
        {"filename": "soc2_type2_gap_analysis.docx", "type": "docx"}
    ]
    
    # Limit number of documents based on size
    num_docs = 3 if size == "small" else (6 if size == "medium" else 12)
    docs_to_seed = doc_templates[:num_docs]
    
    seeded_docs = []
    seeded_reports = []
    
    for idx, template in enumerate(docs_to_seed):
        # Create Document
        doc = Document(
            filename=f"C:/Users/10651.PHNTECHNOLOGY/Desktop/Projects/Enterprise AI/uploads/{1600000000 + idx}_{template['filename']}",
            type=template["type"],
            upload_timestamp=datetime.utcnow() - timedelta(days=num_docs - idx),
            status=WorkflowStatus.PUBLISHED if idx % 2 == 0 else WorkflowStatus.PENDING_REVIEW
        )
        db.add(doc)
        db.commit()
        seeded_docs.append(doc)
        
        # Create Workflow Job
        job = WorkflowJob(
            document_id=doc.id,
            status=WorkflowStatus.PUBLISHED if idx % 2 == 0 else WorkflowStatus.PENDING_REVIEW,
            logs="File uploaded. Saved to uploads.\nText parsed successfully.\nAI analysis complete.\nReports generated.\n",
            updated_at=datetime.utcnow() - timedelta(days=num_docs - idx)
        )
        db.add(job)
        db.commit()
        
        # Create GovernanceReport
        # Set report statuses so we have some in each role's queue:
        # report 0: APPROVED
        # report 1: PENDING_MANAGER_REVIEW
        # report 2: ESCALATED (Governance Lead)
        # report 3: DRAFT with changes_requested (Analyst)
        if idx == 0:
            status_val = "APPROVED"
            rev_status = "approved"
            assigned_to = "Manager"
        elif idx == 1:
            status_val = "PENDING_MANAGER_REVIEW"
            rev_status = "pending_review"
            assigned_to = "Manager"
        elif idx == 2:
            status_val = "ESCALATED"
            rev_status = "pending_review"
            assigned_to = "Governance Lead"
        elif idx == 3:
            status_val = "DRAFT"
            rev_status = "changes_requested"
            assigned_to = "Analyst"
        else:
            status_val = "APPROVED" if idx % 2 == 0 else "PENDING_MANAGER_REVIEW"
            rev_status = "approved" if idx % 2 == 0 else "pending_review"
            assigned_to = "Manager"
            
        report = GovernanceReport(
            document_id=doc.id,
            summary=f"This report covers the governance assessment of {template['filename']}. Key compliance metrics are aligned with enterprise frameworks.",
            executive_summary=f"Executive Summary: {template['filename']} presents a set of findings requiring mitigation.",
            confidence_score=0.85 + (idx % 10) * 0.01,
            model_version="claude-3-5-sonnet",
            prompt_version="v1",
            review_status=rev_status,
            reviewer="reviewer_user" if rev_status == "approved" else None,
            review_notes="Meets organizational standards." if rev_status == "approved" else ("Requires mitigation roadmap clarification" if rev_status == "changes_requested" else None),
            processing_time_seconds=12.4 + idx,
            tokens_used=4200 + idx * 100,
            provider_name="anthropic",
            version=1,
            is_latest=True,
            document_type="governance_report" if idx % 3 == 0 else "project_status_report",
            classification_confidence=0.9,
            governance_relevance="high" if idx % 2 == 0 else "medium",
            created_by="Analyst",
            assigned_to=assigned_to,
            status=status_val,
            created_at=datetime.utcnow() - timedelta(days=num_docs - idx),
            updated_at=datetime.utcnow() - timedelta(days=num_docs - idx)
        )
        db.add(report)
        db.commit()
        seeded_reports.append(report)
        
        # Create AuditLog for Document lifecycle
        audit_uploaded = AuditLog(
            document_id=doc.id,
            event="Uploaded",
            user="analyst_user",
            user_role="Analyst",
            action="document upload",
            entity_type="document",
            entity_id=doc.id,
            details=f"Document uploaded: {template['filename']}",
            timestamp=datetime.utcnow() - timedelta(days=num_docs - idx, hours=1)
        )
        audit_processed = AuditLog(
            document_id=doc.id,
            governance_report_id=report.id,
            event="AI_REPORT_GENERATED",
            user="system",
            user_role="Analyst",
            action="AI report generation",
            entity_type="report",
            entity_id=report.id,
            details="Governance Report generated: V1",
            timestamp=datetime.utcnow() - timedelta(days=num_docs - idx)
        )
        db.add(audit_uploaded)
        db.add(audit_processed)
        db.commit()
        
        if status_val == "APPROVED":
            audit_approved = AuditLog(
                document_id=doc.id,
                governance_report_id=report.id,
                event="REPORT_APPROVED",
                user="manager_user",
                user_role="Manager",
                action="report approval",
                entity_type="report",
                entity_id=report.id,
                details="Governance Report reviewed and approved by Manager.",
                timestamp=datetime.utcnow() - timedelta(days=num_docs - idx, minutes=-30)
            )
            db.add(audit_approved)
            db.commit()
        elif status_val == "ESCALATED":
            audit_escalated = AuditLog(
                document_id=doc.id,
                governance_report_id=report.id,
                event="REPORT_ESCALATED",
                user="manager_user",
                user_role="Manager",
                action="report escalation",
                entity_type="report",
                entity_id=report.id,
                details="Governance Report escalated to Governance Lead due to unresolved high severity risks.",
                timestamp=datetime.utcnow() - timedelta(days=num_docs - idx, minutes=-30)
            )
            db.add(audit_escalated)
            db.commit()
            
    # Now seed RAID items, Escalation items, MitigationTasks, and Notifications
    raid_templates = [
        {
            "type": "risk",
            "description": "Lack of encryption at rest for PII database schema.",
            "severity": "critical",
            "risk_score": 90,
            "owner": "Governance Lead",
            "explain_why": "Leaving PII database columns unencrypted exposes customer data to unauthorized exfiltration and breaches GDPR compliance.",
            "suggested_actions": "1. Enable AES-256 transparent database encryption.\n2. Rotate DB master credentials.\n3. Audit access logs.",
            "estimated_impact": "90% reduction in compliance risk"
        },
        {
            "type": "risk",
            "description": "Default passwords left active on cloud networking endpoints.",
            "severity": "high",
            "risk_score": 80,
            "owner": "Manager",
            "explain_why": "Default credentials on internet-facing cloud endpoints are easily compromised by automated brute-force attacks.",
            "suggested_actions": "1. Enforce password complexity via IAM policy.\n2. Revoke default admin profiles.\n3. Implement MFA.",
            "estimated_impact": "80% reduction in breach probability"
        },
        {
            "type": "issue",
            "description": "Third-party audit vendor delayed security review by two weeks.",
            "severity": "medium",
            "risk_score": 55,
            "owner": "Analyst",
            "explain_why": "Delays in third-party security audits hold up compliance sign-offs and block the deployment of key features.",
            "suggested_actions": "1. Establish backup compliance vendors.\n2. Define vendor penalty terms in SLA.\n3. Escalate delay internally.",
            "estimated_impact": "50% improvement in audit scheduling predictability"
        },
        {
            "type": "dependency",
            "description": "Cloud migration is dependent on external security signoff.",
            "severity": "medium",
            "risk_score": 40,
            "owner": "Manager",
            "explain_why": "Lack of external security signoff blocks migration timelines, risking hardware lease expiration fees.",
            "suggested_actions": "1. Conduct pre-assessment alignment check.\n2. Assign liaison to external signoff team.\n3. Schedule daily progress updates.",
            "estimated_impact": "70% migration schedule slippage risk reduction"
        },
        {
            "type": "risk",
            "description": "Inadequate API rate limiting exposes endpoint to brute force attacks.",
            "severity": "high",
            "risk_score": 75,
            "owner": "Analyst",
            "explain_why": "No API rate limits on authentication routes can lead to credential stuffing and DDoS attacks.",
            "suggested_actions": "1. Deploy rate limiting middleware (max 100 req/min).\n2. Set up WAF rule blocklist.\n3. Enable IP throttling.",
            "estimated_impact": "85% reduction in API brute-force susceptibility"
        },
        {
            "type": "risk",
            "description": "System logging fails to capture administrative privilege changes.",
            "severity": "medium",
            "risk_score": 60,
            "owner": "Manager",
            "explain_why": "Failure to log administrative privilege changes prevents audit validation and compromises insider threat detection.",
            "suggested_actions": "1. Enable cloud audit logging for IAM events.\n2. Stream logs to central SIEM tool.\n3. Configure real-time alerts.",
            "estimated_impact": "75% audit trail visibility enhancement"
        },
        {
            "type": "issue",
            "description": "Critical patch for Kubernetes API container has not been deployed.",
            "severity": "critical",
            "risk_score": 95,
            "owner": "Governance Lead",
            "explain_why": "Unpatched container infrastructure exposes internal cluster networking to remote code execution exploits.",
            "suggested_actions": "1. Apply container patch in staging.\n2. Execute rolling update in production.\n3. Verify cluster version.",
            "estimated_impact": "95% container vulnerability reduction"
        },
        {
            "type": "risk",
            "description": "Weak password policy does not enforce special character requirements.",
            "severity": "low",
            "risk_score": 25,
            "owner": "Analyst",
            "explain_why": "Simple password policies result in weak credentials that are vulnerable to offline dictionary cracking.",
            "suggested_actions": "1. Update IAM password policy parameters.\n2. Force user password resets.\n3. Enable federated SSO.",
            "estimated_impact": "40% credential-security improvement"
        }
    ]
    
    num_raid = 4 if size == "small" else (8 if size == "medium" else 20)
    
    seeded_raid_items = []
    for r_idx in range(num_raid):
        tmpl = raid_templates[r_idx % len(raid_templates)]
        report_to_attach = seeded_reports[r_idx % len(seeded_reports)]
        
        curr_score = tmpl["risk_score"]
        if r_idx % 4 == 0:
            curr_score = int(tmpl["risk_score"] * 0.5)
            
        raid_item = RaidItem(
            report_id=report_to_attach.id,
            type=tmpl["type"],
            description=tmpl["description"],
            severity=tmpl["severity"],
            confidence_score=0.92,
            source_excerpt=f"Identified potential breach area relating to: {tmpl['description']}",
            recommended_mitigations=json.dumps([f"Remediation action for: {tmpl['description']}"]),
            implementation_effort="Medium",
            expected_risk_reduction="High",
            recommended_priority="P1" if tmpl["severity"] in ("critical", "high") else "P2",
            suggested_owner_role=tmpl["owner"],
            priority="P1" if tmpl["severity"] in ("critical", "high") else "P2",
            risk_score=tmpl["risk_score"],
            current_risk_score=curr_score,
            explain_why=tmpl["explain_why"],
            suggested_actions=tmpl["suggested_actions"],
            estimated_impact=tmpl["estimated_impact"],
            explainability_trace=json.dumps({"playbook": "Enterprise Security Controls v4", "matched_keywords": ["pii", "database", "security"], "recommendation_source": "deterministic_rules"})
        )
        db.add(raid_item)
        db.commit()
        seeded_raid_items.append(raid_item)
        
    # Seed Escalations
    esc_templates = [
        {
            "description": "Critical security patching SLA breached for production environment.",
            "severity": "critical",
            "status": "ASSIGNED",
            "owner": "Governance Lead",
            "explain_why": "Unpatched production servers violate SOC2 SLA requirements and expose the network to known public exploits.",
            "suggested_actions": "1. Initiate emergency patching protocol.\n2. Document SLA deviation reason.\n3. Conduct root cause analysis.",
            "estimated_impact": "80% reduction in vulnerability window"
        },
        {
            "description": "Data compliance officer raises concerns over cross-border data transfer policy.",
            "severity": "high",
            "status": "ASSIGNED",
            "owner": "Governance Lead",
            "explain_why": "Cross-border transfer of European resident data without SCC validation violates GDPR regulations, risking major fines.",
            "suggested_actions": "1. Draft standard contractual clauses (SCCs).\n2. Restrict region-specific database replication.\n3. Review compliance audit.",
            "estimated_impact": "95% regulatory compliance improvement"
        },
        {
            "description": "Manager escalates API rate limit issue due to DDoS traffic pattern.",
            "severity": "medium",
            "status": "ASSIGNED",
            "owner": "Manager",
            "explain_why": "Ongoing DDoS traffic pattern consumes API gateway capacity, causing service degradation for legitimate customers.",
            "suggested_actions": "1. Enable cloud security scrubbing center.\n2. Implement CAPTCHA challenge rules.\n3. Set up edge throttling.",
            "estimated_impact": "90% service availability restoration"
        },
        {
            "description": "Steering committee rejects cloud migration risk exceptions.",
            "severity": "critical",
            "status": "OPEN",
            "owner": "Governance Lead",
            "explain_why": "Risk exception rejection means migration is blocked until all identified security controls are fully implemented.",
            "suggested_actions": "1. Revise risk mitigation roadmap.\n2. Implement missing encryption controls.\n3. Resubmit to steering committee.",
            "estimated_impact": "70% control alignment progress"
        },
        {
            "description": "Unassigned security findings overdue by 30+ days.",
            "severity": "high",
            "status": "CLOSED",
            "owner": "Governance Lead",
            "explain_why": "Vulnerabilities left unassigned for over 30 days violate the internal SLA and increase the active exploit window.",
            "suggested_actions": "1. Bulk-assign findings to team leads.\n2. Set mandatory 7-day remediation target.\n3. Report to VP of Security.",
            "estimated_impact": "60% backlog reduction in SLA violations"
        }
    ]
    
    num_esc = 2 if size == "small" else (4 if size == "medium" else 8)
    seeded_esc_items = []
    for e_idx in range(num_esc):
        tmpl = esc_templates[e_idx % len(esc_templates)]
        report_to_attach = seeded_reports[e_idx % len(seeded_reports)]
        
        esc = EscalationItem(
            report_id=report_to_attach.id,
            description=tmpl["description"],
            severity=tmpl["severity"],
            status=tmpl["status"],
            routing_target="Governance Lead" if tmpl["owner"] == "Governance Lead" else "Manager",
            source_excerpt=f"Escalated node in assessment: {tmpl['description']}",
            confidence_score=0.88,
            raised_by="Manager" if tmpl["status"] == "ASSIGNED" else "Analyst",
            assigned_to=tmpl["owner"],
            resolved_by="Governance Lead" if tmpl["status"] == "CLOSED" else None,
            remediation_plan="1. Conduct impact analysis.\n2. Review policy deviation.\n3. Request executive signoff.",
            expected_risk_reduction="High",
            priority="P1" if tmpl["severity"] in ("critical", "high") else "P2",
            suggested_owner_role=tmpl["owner"],
            risk_score=80 if tmpl["severity"] == "critical" else 60,
            explain_why=tmpl["explain_why"],
            suggested_actions=tmpl["suggested_actions"],
            estimated_impact=tmpl["estimated_impact"],
            explainability_trace=json.dumps({"playbook": "Executive Escalation Protocol", "matched_keywords": ["sla", "breach", "critical"], "recommendation_source": "escalation_rules"}),
            created_at=datetime.utcnow() - timedelta(days=num_docs - e_idx)
        )
        db.add(esc)
        db.commit()
        seeded_esc_items.append(esc)
        
    # Seed Mitigation Tasks
    mit_templates = [
        {"title": "Implement TLS 1.3 encryption on database connections", "owner": "Analyst", "status": "IN_PROGRESS", "offset": 14, "pct": 45, "eff": 20},
        {"title": "Configure automated account lockout policy in Active Directory", "owner": "Manager", "status": "PLANNED", "offset": 10, "pct": 0, "eff": 20},
        {"title": "Complete third-party vendor SOC2 Type 2 report review", "owner": "Analyst", "status": "COMPLETED", "offset": 5, "pct": 100, "eff": 15},
        {"title": "Conduct emergency patching cycle for production servers", "owner": "Governance Lead", "status": "IN_PROGRESS", "offset": 12, "pct": 80, "eff": 30},
        {"title": "Remediate SQL injection vulnerability on public auth endpoints", "owner": "Analyst", "status": "VERIFIED", "offset": -2, "pct": 100, "eff": 25},
        {"title": "Run network vulnerability scan across all AWS subnets", "owner": "Manager", "status": "COMPLETED", "offset": -1, "pct": 100, "eff": 15},
        {"title": "Document administrative access policy exceptions", "owner": "Analyst", "status": "IN_PROGRESS", "offset": -3, "pct": 20, "eff": 10},
        {"title": "Verify data backup integrity and replication logs", "owner": "Governance Lead", "status": "IN_PROGRESS", "offset": 1, "pct": 50, "eff": 15}
    ]
    
    num_mit = 4 if size == "small" else (8 if size == "medium" else 15)
    seeded_mit_tasks = []
    for m_idx in range(num_mit):
        tmpl = mit_templates[m_idx % len(mit_templates)]
        raid_to_attach = seeded_raid_items[m_idx % len(seeded_raid_items)]
        
        offset_days = tmpl["offset"]
        # Ensure some tasks are overdue for SLA testing
        if m_idx in (0, 6):
            offset_days = -5
            
        target_date_str = (datetime.utcnow() + timedelta(days=offset_days)).strftime("%Y-%m-%d")
        
        task = MitigationTask(
            title=tmpl["title"],
            description=f"Seeded mitigation task for: {raid_to_attach.description}",
            related_raid_item_id=raid_to_attach.id,
            related_escalation_id=seeded_esc_items[m_idx % len(seeded_esc_items)].id if seeded_esc_items else None,
            owner_role=tmpl["owner"],
            owner_name=f"Demo {tmpl['owner']} User",
            priority="P1" if raid_to_attach.severity in ("critical", "high") else "P2",
            risk_score=raid_to_attach.risk_score,
            target_date=target_date_str,
            sla_status="OVERDUE" if offset_days < 0 and tmpl["status"] not in ("COMPLETED", "VERIFIED") else ("AT_RISK" if 0 <= offset_days <= 3 and tmpl["status"] not in ("COMPLETED", "VERIFIED") else "ON_TRACK"),
            status=tmpl["status"],
            completion_percentage=tmpl["pct"],
            effectiveness=tmpl["eff"],
            created_at=datetime.utcnow() - timedelta(days=10),
            updated_at=datetime.utcnow() - timedelta(days=1),
            completed_at=datetime.utcnow() - timedelta(days=2) if tmpl["status"] in ("COMPLETED", "VERIFIED") else None,
            verified_at=datetime.utcnow() - timedelta(days=1) if tmpl["status"] == "VERIFIED" else None
        )
        db.add(task)
        db.commit()
        seeded_mit_tasks.append(task)
        
        audit_task_create = AuditLog(
            governance_report_id=raid_to_attach.report_id,
            event="MITIGATION_CREATED",
            user="system",
            user_role="Analyst",
            action="mitigation creation",
            entity_type="mitigation",
            entity_id=task.id,
            details=f"Mitigation task '{task.title[:60]}' created.",
            timestamp=datetime.utcnow() - timedelta(days=10)
        )
        db.add(audit_task_create)
        db.commit()
        
        if tmpl["status"] == "VERIFIED":
            audit_task_verify = AuditLog(
                governance_report_id=raid_to_attach.report_id,
                event="MITIGATION_VERIFIED",
                user="gov_lead_user",
                user_role="Governance Lead",
                action="mitigation verification",
                entity_type="mitigation",
                entity_id=task.id,
                details="Mitigation task verified by Governance Lead. Current risk score of related RAID item reduced.",
                timestamp=datetime.utcnow() - timedelta(days=1)
            )
            db.add(audit_task_verify)
            db.commit()
            
    # Seed Notifications
    notif_templates = [
        {"severity": "MEDIUM", "type": "REPORT_PENDING_REVIEW", "title": "Review Required: Q2 Risk Assessment", "msg": "Governance Report for Q2 Risk Assessment is awaiting review.", "role": "Manager"},
        {"severity": "HIGH", "type": "ESCALATION_ASSIGNED", "title": "New Escalation Assigned", "msg": "A critical resource diversion escalation has been assigned to you.", "role": "Governance Lead"},
        {"severity": "MEDIUM", "type": "MITIGATION_ASSIGNED", "title": "Task Assigned: TLS Connection Upgrade", "msg": "A mitigation task for TLS 1.3 migration has been assigned to you.", "role": "Analyst"},
        {"severity": "CRITICAL", "type": "SLA_BREACH", "title": "SLA Breach on Password Rotation Task", "msg": "SLA Breach: Mitigation task to enforce password complexity is 5 days overdue.", "role": "Governance Lead"},
        {"severity": "CRITICAL", "type": "SLA_BREACH", "title": "SLA Breach on Password Rotation Task", "msg": "SLA Breach: Mitigation task to enforce password complexity is 5 days overdue.", "role": "Manager"},
        {"severity": "LOW", "type": "GOVERNANCE_ALERT", "title": "Governance Report Approved", "msg": "Governance report for vendor security audit has been approved.", "role": "Analyst"}
    ]
    
    num_notif = 6 if size == "small" else (12 if size == "medium" else 20)
    for n_idx in range(num_notif):
        tmpl = notif_templates[n_idx % len(notif_templates)]
        related_id = None
        related_type = None
        if tmpl["type"] == "REPORT_PENDING_REVIEW" and seeded_reports:
            related_id = seeded_reports[n_idx % len(seeded_reports)].id
            related_type = "report"
        elif tmpl["type"] == "ESCALATION_ASSIGNED" and seeded_esc_items:
            related_id = seeded_esc_items[n_idx % len(seeded_esc_items)].id
            related_type = "escalation"
        elif tmpl["type"] in ("MITIGATION_ASSIGNED", "SLA_BREACH") and seeded_mit_tasks:
            related_id = seeded_mit_tasks[n_idx % len(seeded_mit_tasks)].id
            related_type = "mitigation"
            
        notif = Notification(
            severity=tmpl["severity"],
            notification_type=tmpl["type"],
            title=tmpl["title"],
            message=tmpl["msg"],
            recipient_role=tmpl["role"],
            related_entity_type=related_type,
            related_entity_id=related_id,
            read_status=(n_idx % 3 == 0),
            created_at=datetime.utcnow() - timedelta(days=num_docs - n_idx)
        )
        db.add(notif)
        db.commit()
        
    for task in db.query(MitigationTask).filter(MitigationTask.status == "VERIFIED").all():
        recalculate_raid_item_risk(db, task.related_raid_item_id)

    # Seed 30 days of snapshots matching the dataset size
    for day in range(30, 0, -1):
        snap_time = datetime.utcnow() - timedelta(days=day)
        progress_factor = (30 - day) / 30.0
        
        health_score = int(60 + progress_factor * 25 + (day % 3) * 2)
        maturity_score = int(55 + progress_factor * 23 + (day % 2) * 1)
        risk_exposure = int(450 - progress_factor * 210 + (day % 5) * 5)
        mitigation_effectiveness_pct = round(40.0 + progress_factor * 35.0 + (day % 4) * 0.5, 1)
        sla_breaches = max(0, int(9 - progress_factor * 8 - (day % 2)))
        open_escalations = max(0, int(5 - progress_factor * 4))
        verified_mitigations = int(progress_factor * 12 + (day % 2))
        critical_risks = max(0, int(6 - progress_factor * 5))
        notification_volume = int(12 + (day % 7) * 4 + int(progress_factor * 10))
        
        snapshot = GovernanceTrendSnapshot(
            timestamp=snap_time,
            health_score=health_score,
            maturity_score=maturity_score,
            risk_exposure=risk_exposure,
            mitigation_effectiveness_pct=mitigation_effectiveness_pct,
            sla_breaches=sla_breaches,
            open_escalations=open_escalations,
            verified_mitigations=verified_mitigations,
            critical_risks=critical_risks,
            notification_volume=notification_volume,
            tenant_id=1
        )
        db.add(snapshot)
    db.commit()
        
    return {"message": f"Successfully generated {size} demo dataset.", "size": size}




from pydantic import BaseModel

class IntegrationSettingsUpdate(BaseModel):
    slack_webhook_url: Optional[str] = None
    teams_webhook_url: Optional[str] = None

@router.get("/governance/integrations", response_model=IntegrationSettingsUpdate)
def get_integrations(
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    org = db.query(Organization).filter(Organization.id == tenant_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organization tenant not found.")
    return {
        "slack_webhook_url": org.slack_webhook_url,
        "teams_webhook_url": org.teams_webhook_url
    }

@router.post("/governance/integrations", response_model=IntegrationSettingsUpdate)
def update_integrations(
    payload: IntegrationSettingsUpdate,
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    org = db.query(Organization).filter(Organization.id == tenant_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organization tenant not found.")
    org.slack_webhook_url = payload.slack_webhook_url
    org.teams_webhook_url = payload.teams_webhook_url
    db.commit()
    db.refresh(org)
    return {
        "slack_webhook_url": org.slack_webhook_url,
        "teams_webhook_url": org.teams_webhook_url
    }

@router.get("/governance/export/csv")
def export_csv(
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    from fastapi.responses import StreamingResponse
    import io
    import csv
    
    raid_items = db.query(RaidItem).filter(RaidItem.tenant_id == tenant_id).all()
    stream = io.StringIO()
    writer = csv.writer(stream)
    
    headers = [
        "ID", "Report ID", "Type", "Description", "Severity", "Risk Score", 
        "Current Score", "Owner Role", "Explain Why", "Suggested Actions", "Estimated Impact"
    ]
    writer.writerow(headers)
    for item in raid_items:
        writer.writerow([
            item.id,
            item.report_id,
            item.type,
            item.description,
            item.severity,
            item.risk_score,
            item.current_risk_score,
            item.suggested_owner_role,
            item.explain_why or "",
            item.suggested_actions or "",
            item.estimated_impact or ""
        ])
    
    response_stream = io.BytesIO(stream.getvalue().encode('utf-8'))
    return StreamingResponse(
        response_stream,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=governance_risk_register.csv"}
    )

@router.get("/governance/export/xlsx")
def export_xlsx(
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant)
):
    from fastapi.responses import StreamingResponse
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    reports = db.query(GovernanceReport).filter(GovernanceReport.tenant_id == tenant_id).all()
    raid_items = db.query(RaidItem).filter(RaidItem.tenant_id == tenant_id).all()
    escalations = db.query(EscalationItem).filter(EscalationItem.tenant_id == tenant_id).all()
    mitigations = db.query(MitigationTask).filter(MitigationTask.tenant_id == tenant_id).all()

    wb = Workbook()
    
    # Sheet 1: Executive Summary
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1["A1"] = "Governance Enterprise Platform"
    ws1["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    ws1["A2"] = f"Executive Governance Summary Report - Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1["A2"].font = Font(name="Calibri", size=11, italic=True, color="595959")
    
    ws1["A4"] = "KPI Metric"
    ws1["B4"] = "Value"
    for cell in ["A4", "B4"]:
        ws1[cell].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws1[cell].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws1[cell].alignment = Alignment(horizontal="left")

    metrics = [
        ("Total Uploaded Documents", len(reports)),
        ("Total Extracted RAID Items", len(raid_items)),
        ("Active Escalations", len(escalations)),
        ("Mitigation Tasks Programmed", len(mitigations))
    ]
    for idx, (m_name, m_val) in enumerate(metrics, start=5):
        ws1[f"A{idx}"] = m_name
        ws1[f"B{idx}"] = m_val
        ws1[f"A{idx}"].font = Font(name="Calibri", size=11, bold=True)
        ws1[f"B{idx}"].font = Font(name="Calibri", size=11)
        if idx % 2 == 1:
            for col in ["A", "B"]:
                ws1[f"{col}{idx}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
    start_row = 11
    ws1.cell(row=start_row, column=1, value="Ingested Documents & Summaries").font = Font(name="Calibri", size=13, bold=True, color="1F4E78")
    
    headers1 = ["Report ID", "Document Name", "Confidence", "Document Type", "Review Status", "Executive Summary"]
    for col_idx, header in enumerate(headers1, start=1):
        cell = ws1.cell(row=start_row+1, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
    for r_idx, rep in enumerate(reports, start=start_row+2):
        doc = db.query(Document).filter(Document.id == rep.document_id).first()
        filename = Path(doc.filename).name if doc else "Unknown"
        ws1.cell(row=r_idx, column=1, value=rep.id)
        ws1.cell(row=r_idx, column=2, value=filename)
        ws1.cell(row=r_idx, column=3, value=rep.confidence_score)
        ws1.cell(row=r_idx, column=4, value=rep.document_type or "governance_report")
        ws1.cell(row=r_idx, column=5, value=rep.review_status)
        ws1.cell(row=r_idx, column=6, value=rep.executive_summary or rep.summary)
        
        if r_idx % 2 == 1:
            for col_idx in range(1, 7):
                ws1.cell(row=r_idx, column=col_idx).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
        
    # Sheet 2: RAID Items
    ws2 = wb.create_sheet(title="RAID Items")
    ws2.views.sheetView[0].showGridLines = True
    headers2 = ["ID", "Report ID", "Type", "Description", "Severity", "Risk Score", "Current Score", "Owner Role", "Explain Why", "Suggested Actions", "Estimated Impact"]
    for col_idx, header in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
    for r_idx, item in enumerate(raid_items, start=2):
        ws2.cell(row=r_idx, column=1, value=item.id)
        ws2.cell(row=r_idx, column=2, value=item.report_id)
        ws2.cell(row=r_idx, column=3, value=item.type)
        ws2.cell(row=r_idx, column=4, value=item.description)
        ws2.cell(row=r_idx, column=5, value=item.severity)
        ws2.cell(row=r_idx, column=6, value=item.risk_score)
        ws2.cell(row=r_idx, column=7, value=item.current_risk_score)
        ws2.cell(row=r_idx, column=8, value=item.suggested_owner_role)
        ws2.cell(row=r_idx, column=9, value=item.explain_why or "")
        ws2.cell(row=r_idx, column=10, value=item.suggested_actions or "")
        ws2.cell(row=r_idx, column=11, value=item.estimated_impact or "")
        
        if r_idx % 2 == 1:
            for col_idx in range(1, len(headers2) + 1):
                ws2.cell(row=r_idx, column=col_idx).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)
        
    # Sheet 3: Escalations
    ws3 = wb.create_sheet(title="Escalations")
    ws3.views.sheetView[0].showGridLines = True
    headers3 = ["ID", "Report ID", "Description", "Severity", "Priority", "Status", "Routing Target", "Assigned To", "Explain Why", "Suggested Actions", "Estimated Impact"]
    for col_idx, header in enumerate(headers3, start=1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
    for r_idx, item in enumerate(escalations, start=2):
        ws3.cell(row=r_idx, column=1, value=item.id)
        ws3.cell(row=r_idx, column=2, value=item.report_id)
        ws3.cell(row=r_idx, column=3, value=item.description)
        ws3.cell(row=r_idx, column=4, value=item.severity)
        ws3.cell(row=r_idx, column=5, value=item.priority)
        ws3.cell(row=r_idx, column=6, value=item.status)
        ws3.cell(row=r_idx, column=7, value=item.routing_target)
        ws3.cell(row=r_idx, column=8, value=item.assigned_to)
        ws3.cell(row=r_idx, column=9, value=item.explain_why or "")
        ws3.cell(row=r_idx, column=10, value=item.suggested_actions or "")
        ws3.cell(row=r_idx, column=11, value=item.estimated_impact or "")
        
        if r_idx % 2 == 1:
            for col_idx in range(1, len(headers3) + 1):
                ws3.cell(row=r_idx, column=col_idx).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)
        
    # Sheet 4: Mitigation Tasks
    ws4 = wb.create_sheet(title="Mitigation Tasks")
    ws4.views.sheetView[0].showGridLines = True
    headers4 = ["ID", "Title", "RAID Item ID", "Owner Role", "Owner Name", "Priority", "Risk Score", "Target Date", "SLA Status", "Status", "Completion %", "Effectiveness"]
    for col_idx, header in enumerate(headers4, start=1):
        cell = ws4.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
    for r_idx, item in enumerate(mitigations, start=2):
        ws4.cell(row=r_idx, column=1, value=item.id)
        ws4.cell(row=r_idx, column=2, value=item.title)
        ws4.cell(row=r_idx, column=3, value=item.related_raid_item_id)
        ws4.cell(row=r_idx, column=4, value=item.owner_role)
        ws4.cell(row=r_idx, column=5, value=item.owner_name)
        ws4.cell(row=r_idx, column=6, value=item.priority)
        ws4.cell(row=r_idx, column=7, value=item.risk_score)
        ws4.cell(row=r_idx, column=8, value=item.target_date)
        ws4.cell(row=r_idx, column=9, value=item.sla_status)
        ws4.cell(row=r_idx, column=10, value=item.status)
        ws4.cell(row=r_idx, column=11, value=item.completion_percentage)
        ws4.cell(row=r_idx, column=12, value=item.effectiveness)
        
        if r_idx % 2 == 1:
            for col_idx in range(1, len(headers4) + 1):
                ws4.cell(row=r_idx, column=col_idx).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
    for col in ws4.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws4.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=governance_risk_register.xlsx"}
    )

